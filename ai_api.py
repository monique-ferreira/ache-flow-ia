# ai_api.py
import os, io, re, time, asyncio
from typing import List, Optional, Dict, Any
from urllib.parse import urlparse, parse_qs, unquote, quote, urljoin
from datetime import datetime, timedelta

import requests
import httpx
import pandas as pd
import pdfplumber
from fastapi import FastAPI, UploadFile, File, Form
from fastapi.responses import JSONResponse
from pydantic import BaseModel

# Vertex (mantido p/ evolução)
from vertexai import init as vertex_init
from vertexai.generative_models import GenerativeModel
from vertexai.preview.language_models import TextEmbeddingModel

# XLSX (preserva hiperlinks)
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

# =========================
# Config
# =========================
PROJECT_ID = os.getenv("GOOGLE_CLOUD_PROJECT")
LOCATION = os.getenv("GOOGLE_CLOUD_LOCATION", "us-central1")
APPLICATION_NAME = os.getenv("GOOGLE_CLOUD_APLICATION", "ai-api")
vertex_init(project=PROJECT_ID, location=LOCATION)

GEMINI_MODEL_ID = os.getenv("GEMINI_MODEL_ID", "gemini-1.5-pro-002")
EMBED_MODEL_ID  = os.getenv("EMBED_MODEL_ID",  "text-embedding-004")

# ---- Sua API no Render ----
TASKS_API_BASE           = os.getenv("TASKS_API_BASE", "https://ache-flow-back.onrender.com").rstrip("/")
TASKS_API_PROJECTS_PATH  = os.getenv("TASKS_API_PROJECTS_PATH", "/projetos")
TASKS_API_TASKS_PATH     = os.getenv("TASKS_API_TASKS_PATH", "/tarefas")
TASKS_API_TOKEN_PATH     = os.getenv("TASKS_API_TOKEN_PATH", "/token")
TASKS_API_USERNAME       = os.getenv("TASKS_API_USERNAME")  # p/ /token
TASKS_API_PASSWORD       = os.getenv("TASKS_API_PASSWORD")  # p/ /token

TIMEOUT_S = int(os.getenv("TIMEOUT_S", "90"))
GENERIC_USER_AGENT = os.getenv("GENERIC_USER_AGENT", "ache-flow-ia/1.0 (+https://tistto.com.br)")
PDF_USER_AGENT     = os.getenv("PDF_USER_AGENT", GENERIC_USER_AGENT)

app = FastAPI(title=f"{APPLICATION_NAME} (Vertex AI)", version="1.4.0")

# =========================
# Handler global — nunca mais “null”
# =========================
@app.exception_handler(Exception)
async def all_exception_handler(request, exc):
    import traceback
    tb = traceback.format_exc()
    return JSONResponse(
        status_code=500,
        content={"error": "internal_error", "detail": str(exc), "trace": tb[-4000:]},
    )

# =========================
# Vertex utils (RAG simples – mantido)
# =========================
def gemini_generate(prompt: str, temperature: float = 0.2) -> str:
    model = GenerativeModel(GEMINI_MODEL_ID)
    response = model.generate_content(prompt, generation_config={"temperature": temperature})
    return response.text or ""

def embed_texts(texts: List[str]) -> List[List[float]]:
    model = TextEmbeddingModel.from_pretrained(EMBED_MODEL_ID)
    embeddings = model.get_embeddings(texts)
    return [e.values for e in embeddings]

# =========================
# Helpers de download (XLSX e PDF)
# =========================
def fetch_bytes(url: str) -> bytes:
    if not url:
        raise ValueError("URL ausente")
    u = url.strip()
    if not u.lower().startswith(("http://", "https://")):
        raise ValueError("URL inválida")
    r = requests.get(u, timeout=TIMEOUT_S, allow_redirects=True, headers={"User-Agent": GENERIC_USER_AGENT})
    r.raise_for_status()
    return r.content

def _attempts_summary(attempts):
    out = []
    for a in attempts:
        out.append({
            "tried": a.get("tried"),
            "status": a.get("status"),
            "content_type": a.get("content_type"),
            "final_url": a.get("final_url"),
            "is_pdf_signature": a.get("is_pdf_signature"),
            "size_kb": (a.get("size") or 0) // 1024
        })
    return out

def _try_download_traced(url: str, timeout: int, user_agent: str):
    # Não use r.raw.read(); deixe o requests tratar chunked/decoding sozinho
    headers = {
        "User-Agent": user_agent,
        # dica para o servidor nos dar bytes "normais"
        "Accept": "application/pdf, */*",
    }
    with requests.get(
        url,
        timeout=timeout,
        allow_redirects=True,
        headers=headers,
        stream=False,  # <— importante
    ) as r:
        status = r.status_code
        ctype = (r.headers.get("Content-Type") or "").lower()
        final_url = r.url
        content = r.content or b""
        return {
            "status": status,
            "content_type": ctype,
            "final_url": final_url,
            "content": content,
            "is_pdf_signature": content[:4] == b"%PDF",
            "size": len(content),
        }

def normalize_sharepoint_pdf_url(u: str) -> str:
    """
    onedrive.aspx?id=...  ->  https://host/<path-encodado-certo>
    (sem ?download=1 aqui; isso entra como fallback no fetch)
    """
    try:
        pu = urlparse(u)
        if pu.netloc and "sharepoint.com" in pu.netloc and pu.path.endswith("/onedrive.aspx"):
            qs = parse_qs(pu.query or "")
            idp = qs.get("id", [None])[0]
            if idp:
                raw_path = unquote(idp)  # /personal/.../Documents/.../Dataset+Link/Doc.Teste.pdf
                if not raw_path.startswith("/"):
                    raw_path = "/" + raw_path
                encoded_path = quote(raw_path, safe="/-_.()~")
                return f"{pu.scheme}://{pu.netloc}{encoded_path}"
    except Exception:
        pass
    return u

def fetch_pdf_bytes(url: str, collect_attempts: bool = False):
    if not url:
        raise ValueError("URL ausente para PDF")
    u = url.strip()
    if not u.lower().startswith(("http://", "https://")):
        raise ValueError("URL inválida para PDF")

    # normaliza 'onedrive.aspx?id=...' -> link direto re-encodado
    u_norm = normalize_sharepoint_pdf_url(u)

    def _is_pdf(ctype: str, content: bytes) -> bool:
        return ("pdf" in ctype) or (content[:4] == b"%PDF")

    attempts = []
    variants = [("direct", u_norm)]

    if "download=1" not in u_norm.lower():
        sep = "&" if "?" in u_norm else "?"
        variants.append(("download=1", u_norm + f"{sep}download=1"))

    if "sharepoint.com" in u_norm.lower():
        pu = urlparse(u_norm)
        base = f"{pu.scheme}://{pu.netloc}"
        src = quote(u_norm, safe="")  # URL inteiro encodado
        variants.append(("download.aspx", f"{base}/_layouts/15/download.aspx?SourceUrl={src}"))

    last = None
    for label, cand in variants:
        r = _try_download_traced(cand, TIMEOUT_S, PDF_USER_AGENT)
        r["tried"] = label
        attempts.append(r)
        last = r
        if r["status"] == 200 and _is_pdf(r["content_type"], r["content"]):
            return (r["content"], attempts) if collect_attempts else r["content"]

    msg = f"Não foi possível obter PDF (último status={last['status'] if last else None}, " \
          f"content-type={last['content_type'] if last else None}, final_url={last['final_url'] if last else None})."
    if collect_attempts:
        raise RuntimeError(msg + " ::attempts:: " + repr(_attempts_summary(attempts)))
    raise ValueError(msg)

def clean_pdf_text(s: str) -> str:
    if not s:
        return s
    # junta linhas quebradas desnecessárias
    s = re.sub(r"[ \t]*\n[ \t]*", " ", s)
    # colapsa espaços múltiplos
    s = re.sub(r"\s{2,}", " ", s)
    # remove espaços antes de pontuação
    s = re.sub(r"\s+([,;\.\!\?\:\)])", r"\1", s)
    # normaliza espaço depois de pontuação
    s = re.sub(r"([,;\.\!\?\:])([^\s])", r"\1 \2", s)
    return s.strip()

# =========================
# PDF: âncoras Texto.1 / Texto1 / Texto4.
# =========================
ANCHOR_VARIANTS = r"(?:Texto\.?\d+)\.?"

def _anchor_regex_flex(label: str) -> re.Pattern:
    m = re.search(r"(?i)texto\.?(\d+)", label or "")
    if not m:
        return re.compile(r"(?!)")
    num = m.group(1)
    return re.compile(rf"(?i)\bTexto\.?{re.escape(num)}\.?\b[:\-]?\s*")

def extract_after_anchor_from_pdf(pdf_bytes: bytes, anchor_label: str, max_chars: int = 4000) -> str:
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        text_all = "\n".join((page.extract_text() or "") for page in pdf.pages)
    if not text_all.strip():
        # PDF sem camada de texto (provável imagem/scan)
        return ""
    rx = _anchor_regex_flex(anchor_label)
    m = rx.search(text_all)
    if not m:
        return ""
    start = m.end()
    next_m = re.search(r"(?i)\bTexto\.?\d+\.?\b", text_all[start:])
    end = start + next_m.start() if next_m else len(text_all)
    return text_all[start:end].strip()[:max_chars].strip()

# =========================
# XLSX → DataFrame preservando hiperlinks da célula
# =========================
def xlsx_bytes_to_dataframe_preserving_hyperlinks(xlsx_bytes: bytes) -> pd.DataFrame:
    """
    Lê a primeira planilha:
    - cabeçalhos na primeira linha
    - valores de célula como texto
    - SE a coluna "Documento Referência" tiver hyperlink, usa o URL real (cell.hyperlink.target)
    Compatível mesmo quando hyperlinks estão como objetos separados.
    """
    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=False)
    ws = wb.active  # primeira aba

    # Cabeçalhos
    headers: List[str] = []
    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=False))
    for cell in first_row:
        headers.append(str(cell.value).strip() if cell.value is not None else "")

    # Mapa de hyperlinks
    hyperlink_map: Dict[str, str] = {}
    for hl in getattr(ws, "_hyperlinks", []) or []:
        try:
            ref = getattr(hl, "ref", None)
            target = getattr(hl, "target", None) or getattr(hl, "location", None)
            if not ref or not target:
                continue
            if ":" in ref:
                min_col, min_row, max_col, max_row = range_boundaries(ref)
                for r in range(min_row, max_row + 1):
                    for c in range(min_col, max_col + 1):
                        coord = ws.cell(row=r, column=c).coordinate
                        hyperlink_map[coord] = target
            else:
                hyperlink_map[ref] = target
        except Exception:
            continue

    rows: List[Dict[str, Any]] = []
    try:
        col_idx_doc = headers.index("Documento Referência")
    except ValueError:
        col_idx_doc = -1

    for row in ws.iter_rows(min_row=2, values_only=False):
        if all((c.value is None or str(c.value).strip() == "") for c in row):
            continue

        record: Dict[str, Any] = {}
        for i, cell in enumerate(row):
            header = headers[i] if i < len(headers) else f"col{i+1}"
            val = cell.value

            if i == col_idx_doc:
                url = None
                if getattr(cell, "hyperlink", None) and getattr(cell.hyperlink, "target", None):
                    url = cell.hyperlink.target
                if not url:
                    url = hyperlink_map.get(cell.coordinate)
                if not url and isinstance(val, str) and val.strip().lower().startswith(("http://", "https://")):
                    url = val.strip()
                record[header] = url or (val if val is not None else "")
            else:
                record[header] = val if val is not None else ""

        rows.append(record)

    return pd.DataFrame(rows)

# =========================
# Auth: obter e cachear token do /token
# =========================
_token_cache: Dict[str, Any] = {"access_token": None, "expires_at": 0, "user_id": None}

async def get_auth_header(client: httpx.AsyncClient) -> Dict[str, str]:
    bearer_env = os.getenv("TASKS_API_BEARER")
    if bearer_env:
        return {"Authorization": f"Bearer {bearer_env}"}

    now = time.time()
    if _token_cache.get("access_token") and now < _token_cache.get("expires_at", 0) - 30:
        return {"Authorization": f"Bearer {_token_cache['access_token']}"}

    if not TASKS_API_USERNAME or not TASKS_API_PASSWORD:
        return {}

    token_url = urljoin(TASKS_API_BASE + "/", TASKS_API_TOKEN_PATH.lstrip("/"))
    data = {"username": TASKS_API_USERNAME, "password": TASKS_API_PASSWORD}
    headers = {"Content-Type": "application/x-www-form-urlencoded"}

    for attempt in range(3):
        try:
            resp = await client.post(token_url, data=data, headers=headers, timeout=TIMEOUT_S)
            if resp.status_code >= 500:
                raise httpx.TransportError(f"server {resp.status_code}")
            resp.raise_for_status()
            payload = resp.json()
            access_token = payload.get("access_token") or payload.get("token") or payload.get("accessToken")
            expires_in  = int(payload.get("expires_in") or 3600)
            _token_cache["user_id"] = payload.get("id") or _token_cache.get("user_id")
            if not access_token:
                raise RuntimeError(f"Resposta de token sem access_token: {payload}")
            _token_cache["access_token"] = access_token
            _token_cache["expires_at"] = time.time() + expires_in
            return {"Authorization": f"Bearer {access_token}"}
        except Exception:
            if attempt == 2:
                raise
            await asyncio.sleep(1.5 * (attempt + 1))
    return {}

# =========================
# Tipos (tarefas)
# =========================
class CreateTaskItem(BaseModel):
    titulo: str
    descricao: Optional[str] = None
    responsavel: Optional[str] = None   # nome/email da planilha
    deadline: Optional[str] = None      # ex.: '5 dias'
    doc_ref: Optional[str] = None       # link PDF p/ documento_referencia
    prazo_data: Optional[str] = None    # YYYY-MM-DD (se houver coluna com data)

async def create_task(client: httpx.AsyncClient, projeto_id: str, responsavel_id: str, item: CreateTaskItem) -> Dict[str, Any]:
    url = f"{TASKS_API_BASE}{TASKS_API_TASKS_PATH}"
    auth = await get_auth_header(client)
    prazo = (item.prazo_data or "").strip() or duration_to_date(item.deadline)

    payload = {
        "nome": item.titulo,
        "descricao": item.descricao,
        "projeto_id": projeto_id,
        "responsavel_id": responsavel_id,
        "prazo": prazo,
        "documento_referencia": item.doc_ref,
        # prioridade/status têm defaults no schema
    }

    r = await client.post(url, json=payload, headers={**auth, "Content-Type": "application/json"}, timeout=TIMEOUT_S)
    r.raise_for_status()
    return r.json()

# =========================
# Integração com API tradicional
# =========================
async def ensure_project_exists_by_id(client: httpx.AsyncClient, projeto_id: str) -> bool:
    url = f"{TASKS_API_BASE}{TASKS_API_PROJECTS_PATH}/{projeto_id}"
    auth = await get_auth_header(client)
    r = await client.get(url, headers=auth, timeout=TIMEOUT_S)
    return r.status_code == 200

async def find_project_id_by_name(client: httpx.AsyncClient, projeto_nome: str) -> Optional[str]:
    url = f"{TASKS_API_BASE}{TASKS_API_PROJECTS_PATH}"
    auth = await get_auth_header(client)
    r = await client.get(url, headers=auth, timeout=TIMEOUT_S)
    if r.status_code != 200:
        return None
    try:
        items = r.json()
        if isinstance(items, list):
            hit = next((p for p in items if str(p.get("nome")) == projeto_nome), None)
            return (hit or {}).get("_id") if hit else None
    except Exception:
        return None
    return None
    
async def create_project_api(
    client: httpx.AsyncClient,
    nome: str,
    responsavel_id: str,
    situacao: str,
    prazo_yyyy_mm_dd: str,
    descricao: Optional[str] = None,
    categoria: Optional[str] = None,
) -> Dict[str, Any]:
    """
    POST /projetos — ProjetoCreate exige: nome, responsavel_id, situacao, prazo (date).
    """
    url = f"{TASKS_API_BASE}{TASKS_API_PROJECTS_PATH}"
    auth = await get_auth_header(client)
    payload = {
        "nome": nome,
        "responsavel_id": responsavel_id,
        "situacao": situacao,
        "prazo": prazo_yyyy_mm_dd,
    }
    if descricao:
        payload["descricao"] = descricao
    if categoria:
        payload["categoria"] = categoria

    r = await client.post(url, json=payload, headers={**auth, "Content-Type": "application/json"}, timeout=TIMEOUT_S)
    r.raise_for_status()
    return r.json()


async def list_funcionarios(client: httpx.AsyncClient) -> List[Dict[str, Any]]:
    url = f"{TASKS_API_BASE}/funcionarios"
    auth = await get_auth_header(client)
    r = await client.get(url, headers=auth, timeout=TIMEOUT_S)
    if r.status_code != 200:
        return []
    try:
        data = r.json()
        return data if isinstance(data, list) else []
    except Exception:
        return []

async def resolve_responsavel_id(client: httpx.AsyncClient, nome_ou_email: Optional[str]) -> Optional[str]:
    """
    Tenta casar por nome/email. Se não achar, usa o id que veio no /token (cache).
    """
    nome_ou_email = (nome_ou_email or "").strip()
    if not nome_ou_email:
        return _token_cache.get("user_id")

    pessoas = await list_funcionarios(client)
    key = nome_ou_email.lower()

    # match por email
    for p in pessoas:
        email = str(p.get("email") or "").lower()
        if email and email == key:
            return p.get("_id")

    # match por nome "exato" (case-insensitive)
    for p in pessoas:
        nome = str(p.get("nome") or "").lower()
        sobrenome = str(p.get("sobrenome") or "").lower()
        full = f"{nome} {sobrenome}".strip()
        if full == key or nome == key:
            return p.get("_id")

    # fallback no próprio usuário do token
    return _token_cache.get("user_id")

def duration_to_date(duracao: Optional[str]) -> str:
    """
    Converte '5 dias' -> (hoje + 5). Retorna YYYY-MM-DD.
    Se vazio/inesperado, usa hoje + 7.
    """
    base = datetime.utcnow().date()
    try:
        s = (duracao or "").strip().lower()
        m = re.search(r"(\d+)", s)
        n = int(m.group(1)) if m else 7
    except Exception:
        n = 7
    return (base + timedelta(days=n)).isoformat()

# =========================
# Endpoint: importar tarefas via .xlsx
# =========================
@app.post("/tasks/from-xlsx")
async def tasks_from_xlsx(
    projeto_id: Optional[str] = Form(None),
    projeto_nome: Optional[str] = Form(None),

    # NOVO: criação de projeto (se não existir)
    create_project_flag: int = Form(0),                 # 1 = cria se não achar
    projeto_situacao: Optional[str] = Form(None),  # obrigatório se for criar
    projeto_prazo: Optional[str] = Form(None),     # YYYY-MM-DD (ou vazio, usa duração padrão)
    projeto_responsavel: Optional[str] = Form(None),  # nome/email p/ resolver responsável do projeto
    projeto_descricao: Optional[str] = Form(None),
    projeto_categoria: Optional[str] = Form(None),

    # FONTE da planilha
    xlsx_url: Optional[str] = Form(None),
    file: Optional[UploadFile] = File(None),

    # debug
    debug: int = Form(0),
):
    # normaliza xlsx_url
    if xlsx_url:
        xlsx_url = xlsx_url.strip()
        if not xlsx_url.lower().startswith(("http://", "https://")):
            xlsx_url = None

    if not xlsx_url and not file:
        return {"erro": "Forneça 'xlsx_url' (http/https) OU envie um arquivo .xlsx"}

    # carrega planilha preservando hiperlinks
    try:
        if xlsx_url:
            xbytes = fetch_bytes(xlsx_url)
            df = xlsx_bytes_to_dataframe_preserving_hyperlinks(xbytes)
        else:
            content = await file.read()
            df = xlsx_bytes_to_dataframe_preserving_hyperlinks(content)
    except Exception as e:
        return {"erro": f"Falha ao ler a planilha (com hiperlinks): {type(e).__name__}: {e}"}

    # colunas esperadas
    required = {"Nome", "Como Fazer", "Documento Referência"}
    missing = required - set(df.columns)
    if missing:
        return {"erro": f"Colunas faltando: {', '.join(missing)}"}

    # função: resolver “Como Fazer?” a partir do PDF indicado no hyperlink
    def resolve_descricao(row) -> str:
        como  = str(row.get("Como Fazer") or "").strip()
        docrf = str(row.get("Documento Referência") or "").strip()
        if not como or not docrf:
            return como

        token_pattern = re.compile(r"(?i)\b((?:Doc\.?\s*)?(Texto\.?\d+))\b\.?")
        if not token_pattern.search(como):
            return como

        try:
            if debug:
                pdf_bytes, _attempts = fetch_pdf_bytes(docrf, collect_attempts=True)
            else:
                pdf_bytes = fetch_pdf_bytes(docrf)
        except Exception:
            return como

        def _repl(m: re.Match) -> str:
            full_token = m.group(1)
            anchor     = m.group(2)
            extracted  = extract_after_anchor_from_pdf(pdf_bytes, anchor)
            if extracted:
                extracted = clean_pdf_text(extracted)
            return extracted if extracted else full_token

        result = token_pattern.sub(_repl, como)
        if debug and result == como:
            return f"{result} [NO-REPLACE]"
        return result

    # aplica a resolução
    df = df.copy()
    df["descricao_final"] = df.apply(resolve_descricao, axis=1)

    # preview
    preview: List[Dict[str, Any]] = []
    first_trace = None

    # coleta trace da primeira linha (se debug=1)
    if debug and len(df) > 0:
        try:
            first_href = str(df.iloc[0].get("Documento Referência") or "").strip()
            _, attempts = fetch_pdf_bytes(first_href, collect_attempts=True)
            first_trace = _attempts_summary(attempts)
        except Exception as e:
            first_trace = f"error: {e}"

    for _, row in df.iterrows():
        titulo = str(row["Nome"])
        descricao_final = str(row.get("descricao_final") or "")
        responsavel_txt = str(row.get("Responsavel") or "")
        duracao_txt = str(row.get("Duração") or "")
        doc_ref = str(row.get("Documento Referência") or "").strip()

        # se existir coluna 'Prazo' já em data (YYYY-MM-DD), use
        prazo_col = str(row.get("Prazo") or "").strip()
        if prazo_col:
            # tenta normalizar p/ YYYY-MM-DD
            try:
                # aceita dd/mm/yyyy também
                if re.match(r"^\d{2}/\d{2}/\d{4}$", prazo_col):
                    d, m, y = prazo_col.split("/")
                    prazo_col = f"{y}-{m}-{d}"
            except Exception:
                pass

        preview.append({
            "titulo": titulo,
            "descricao": descricao_final,
            "responsavel": responsavel_txt,
            "deadline": duracao_txt,
            "doc_ref": doc_ref,
            "prazo": prazo_col or None,
        })
            
    # se não veio projeto → só preview
    if projeto_id is None and not projeto_nome:
        resp = {
            "mode": "preview",
            "tarefas": preview[:50],
            "total": len(preview),
            "instrucoes": "Para atribuir, chame novamente passando 'projeto_id' (ou 'projeto_nome' já existente)."
        }
        if debug:
            resp["_pdf_trace_first_row"] = first_trace
        return resp

    # com projeto → atribui
    async with httpx.AsyncClient() as client:
        # 1) resolver projeto
        resolved_project_id: Optional[str] = projeto_id
        if not resolved_project_id and projeto_nome:
            resolved_project_id = await find_project_id_by_name(client, projeto_nome)

        if not resolved_project_id:
            if create_project_flag:
                # resolver responsavel do PROJETO; se não vier, usa o usuário do token
                proj_resp_id = await resolve_responsavel_id(client, projeto_responsavel)
                # prazo do projeto: se não vier data, usa hoje + 30 (exemplo)
                proj_prazo = (projeto_prazo or "").strip() or (datetime.utcnow().date() + timedelta(days=30)).isoformat()
                # situacao: se não vier, define 'em andamento' (ou outro estado da sua régua)
                situacao = (projeto_situacao or "em andamento").strip()
                # cria
                proj = await create_project_api(
                    client=client,
                    nome=projeto_nome or "Projeto sem nome",
                    responsavel_id=proj_resp_id,
                    situacao=situacao,
                    prazo_yyyy_mm_dd=proj_prazo,
                    descricao=projeto_descricao,
                    categoria=projeto_categoria,
                )
                resolved_project_id = proj.get("_id") or proj.get("id")
            else:
                return {"erro": "Projeto não encontrado. Informe 'projeto_id' (ou 'projeto_nome' existente) "
                                "ou envie 'create_project_flag=1' + 'projeto_situacao' + 'projeto_prazo'."}

        # dupla confirmação por ID
        ok = await ensure_project_exists_by_id(client, resolved_project_id)
        if not ok:
            return {"erro": f"Projeto {resolved_project_id} não encontrado na API tradicional."}

        # 2) criar tarefas
        created = []
        for item in preview:
            resp_id = await resolve_responsavel_id(client, item.get("responsavel"))
            payload = CreateTaskItem(
                titulo=item["titulo"],
                descricao=item["descricao"],
                responsavel=item.get("responsavel") or None,
                deadline=item.get("deadline") or None,
                doc_ref=item.get("doc_ref") or None,
                prazo_data=item.get("prazo") or None,
            )
            try:
                created.append(await create_task(client, resolved_project_id, resp_id, payload))
            except Exception as e:
                created.append({"erro": str(e), "titulo": item["titulo"]})

    resp = {"mode": "assigned", "projeto_id": resolved_project_id, "criados": created, "total": len(created)}
    if debug:
        resp["_pdf_trace_first_row"] = first_trace
    return resp

# =========================
# Root
# =========================
@app.get("/")
def root():
    return {"status": "AI API rodando com Vertex AI", "project": PROJECT_ID, "location": LOCATION}

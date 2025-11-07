import os, io, re, time, asyncio, json
from typing import List, Optional, Dict, Any, Tuple
from collections.abc import Mapping
from urllib.parse import urlparse, parse_qs, unquote, quote, urljoin
from datetime import datetime, timedelta, date

import requests
import httpx
import pandas as pd
import pdfplumber
import fitz # PyMuPDF
from fastapi import FastAPI, UploadFile, File, Form, Header, HTTPException, Depends
from fastapi.responses import JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, Field

# Vertex AI
from vertexai import init as vertex_init
from vertexai.generative_models import (
    GenerativeModel,
    Tool,
    FunctionDeclaration,
    Part,
    Content,
)

# Mongo
from pymongo import MongoClient, ReturnDocument
from bson import ObjectId, DBRef

# XLSX
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

from dotenv import load_dotenv
load_dotenv()

class MessageContent(BaseModel):
    tipo_resposta: str
    conteudo_texto: str
    dados: Optional[List[Any]] = None

class HistoryMessage(BaseModel):
    sender: str
    content: MessageContent

class ChatRequest(BaseModel):
    pergunta: str
    history: Optional[List[HistoryMessage]] = None
    nome_usuario: Optional[str] = None
    email_usuario: Optional[str] = None
    id_usuario: Optional[str] = None

class ChatWithPdfUrlRequest(BaseModel):
    pergunta: str
    pdf_url: str = Field(..., description="A URL completa para o arquivo PDF (incluindo Google Drive, Sharepoint, etc.)")
    nome_usuario: Optional[str] = None
    email_usuario: Optional[str] = None
    id_usuario: Optional[str] = None

# =========================
# Config
# =========================
PROJECT_ID = os.getenv("GOOGLE_CLOUD_PROJECT")
LOCATION = os.getenv("GOOGLE_CLOUD_LOCATION", "us-central1")
APPLICATION_NAME = os.getenv("GOOGLE_CLOUD_APLICATION", "ai-service")
GEMINI_MODEL_ID = os.getenv("GEMINI_MODEL_ID", "gemini-2.0-flash-001")

API_KEY = os.getenv("API_KEY") 
MONGO_URI = os.getenv("MONGO_URI", "mongodb://localhost:27017/acheflow")
COLL_PROJETOS = os.getenv("MONGO_COLL_PROJETOS", "projetos")
COLL_TAREFAS = os.getenv("MONGO_COLL_TAREFAS", "tarefas")
COLL_FUNCIONARIOS = os.getenv("MONGO_COLL_FUNCIONARIOS", "funcionarios")
COLL_CONTEXTOS = os.getenv("MONGO_COLL_CONTEXTOS", "chat_contexts")

TASKS_API_BASE           = os.getenv("TASKS_API_BASE", "https://ache-flow-back.onrender.com").rstrip("/")
TASKS_API_PROJECTS_PATH  = os.getenv("TASKS_API_PROJECTS_PATH", "/projetos")
TASKS_API_TASKS_PATH     = os.getenv("TASKS_API_TASKS_PATH", "/tarefas")
TASKS_API_TOKEN_PATH     = os.getenv("TASKS_API_TOKEN_PATH", "/token")
TASKS_API_USERNAME       = os.getenv("TASKS_API_USERNAME")
TASKS_API_PASSWORD       = os.getenv("TASKS_API_PASSWORD")
ACHEFLOW_MAIN_API_TOKEN = os.getenv("ACHEFLOW_API_TOKEN") 

TIMEOUT_S = 90
GENERIC_USER_AGENT = "ache-flow-ia/1.0"
PDF_USER_AGENT     = GENERIC_USER_AGENT

MAX_TOOL_STEPS = 6
DEFAULT_TOP_K = 8

# =========================
# FastAPI App (Único)
# =========================
app = FastAPI(title=f"{APPLICATION_NAME} (Serviço Unificado de IA e Importação)", version="19.0.0")

origins = [
    "http://localhost:5173",
    "http://localhost:5174",
    "https://ache-flow.vercel.app",
    "https://www.ache-flow.vercel.app"
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)
# ===============================

# =========================
# Segurança
# =========================
def require_api_key(x_api_key: Optional[str] = Header(None)):
    if API_KEY and (x_api_key or "") != API_KEY:
        raise HTTPException(status_code=401, detail="Invalid API Key")

# =========================
# Handler de Erro
# =========================
@app.exception_handler(Exception)
async def all_exception_handler(request, exc):
    import traceback
    tb = traceback.format_exc()
    detail = str(exc)
    status_code = 500
    
    if isinstance(exc, httpx.HTTPStatusError):
        try:
            detail = exc.response.json().get("detail", str(exc))
            status_code = exc.response.status_code
        except Exception:
            detail = exc.response.text
    elif isinstance(exc, HTTPException):
        detail = exc.detail
        status_code = exc.status_code

    trace_info = tb[-4000:] if "localhost" in str(request.url) or os.getenv("ENV_MODE") == "debug" else None
    return JSONResponse(
        status_code=status_code,
        content={"erro": "internal_error", "detail": detail, "trace": trace_info},
    )

# =========================
# Helpers (Datas, Mongo, etc)
# =========================
def today() -> datetime: return datetime.utcnow()
def iso_date(d: datetime) -> str: return d.date().isoformat()

def month_bounds(d: datetime) -> Tuple[str, str]:
    first = d.replace(day=1).date().isoformat()
    if d.month == 12:
        nxt = d.replace(year=d.year + 1, month=1, day=1)
    else:
        nxt = d.replace(month=d.month + 1, day=1)
    last = (nxt - timedelta(days=1)).date().isoformat()
    return first, last

def to_oid(id_str: str) -> ObjectId:
    try: return ObjectId(id_str)
    except Exception: return id_str

def pick(d: Dict[str, Any], keys: List[str]) -> Dict[str, Any]:
    return {k: d.get(k) for k in keys if k in d}

def extract_gsheet_id(url: str) -> Optional[str]:
    if not url: return None
    m = re.search(r"/d/([a-zA-Z0-9_-]{40,})", url)
    if m:
        return m.group(1)
    return None

def sanitize_doc(data: Any) -> Any:
    if isinstance(data, (datetime, date)):
        return data.isoformat()
    if isinstance(data, ObjectId):
        return str(data)
    if isinstance(data, DBRef):
        return str(data.id)
    data_type_name = data.__class__.__name__
    if isinstance(data, (Mapping, dict)) or data_type_name == 'MapComposite':
        if len(data) == 1 and (list(data.keys())[0].startswith('$') or list(data.keys())[0] == '_id'):
            return sanitize_doc(list(data.values())[0])
        return {k: sanitize_doc(v) for k, v in data.items()}
    if isinstance(data, list):
        return [sanitize_doc(item) for item in data]
    return data

def mongo():
    if not MONGO_URI: 
        raise RuntimeError("MONGO_URI não foi definida")
    client = MongoClient(MONGO_URI)
    try:
        db = client.get_default_database()
        db.command("ping") 
        return db
    except Exception as e:
        raise RuntimeError(f"Não foi possível conectar ao MongoDB. Verifique a MONGO_URI e o firewall do Atlas. Erro: {e}")

def _parse_date_robust(date_str: str) -> str:
    date_str = (date_str or "").strip()
    try:
        return datetime.strptime(date_str, "%d/%m/%Y").date().isoformat()
    except ValueError:
        try:
            return datetime.strptime(date_str, "%d-%m-%Y").date().isoformat()
        except ValueError:
            return date_str

def _get_employee_map() -> Dict[str, str]:
    try:
        employees_raw = mongo()[COLL_FUNCIONARIOS].find({}, {"nome": 1, "sobrenome": 1, "_id": 1})
        employees = [sanitize_doc(x) for x in employees_raw]
        return {
            str(emp.get("_id")): f"{emp.get('nome', '')} {emp.get('sobrenome', '')}".strip()
            for emp in employees
            if emp.get("_id")
        }
    except Exception as e:
        print(f"Erro ao buscar mapa de funcionários: {e}")
        return {}

def _enrich_doc_with_responsavel(doc: Dict[str, Any], employee_map: Dict[str, str]) -> Dict[str, Any]:
    resp_id_key = None
    if "responsavel" in doc:
        resp_id_key = "responsavel"
    elif "responsavel_id" in doc:
        resp_id_key = "responsavel_id"
    else:
        doc["responsavel_nome"] = "(Responsável não atribuído)"
        return doc

    resp_id = str(doc.get(resp_id_key)) 
    
    if resp_id and resp_id != "None":
        if resp_id in employee_map:
            doc["responsavel_nome"] = employee_map[resp_id]
        else:
            doc["responsavel_nome"] = f"(ID não encontrado: {resp_id})"
    else:
        doc["responsavel_nome"] = "(Responsável não atribuído)"

    if resp_id_key in doc:
        del doc[resp_id_key]
        
    return doc

def _find_task_id_by_name_sync(task_name: str, project_name: str) -> Optional[str]:
    """
    Helper interno para encontrar o _id de uma tarefa com base no nome
    e no nome do projeto (para desambiguação).
    """
    project_name_norm = (project_name or "").strip()
    task_name_norm = (task_name or "").strip()
    if not project_name_norm or not task_name_norm:
        return None
    try:
        # 1. Encontrar o Projeto
        rx_proj = {"$regex": f"^{re.escape(project_name_norm)}$", "$options": "i"}
        proj = mongo()[COLL_PROJETOS].find_one({"nome": rx_proj}, {"_id": 1})
        if not proj: 
            print(f"[Helper] Projeto '{project_name_norm}' não encontrado.")
            return None
        project_id = proj.get("_id")
        
        # 2. Encontrar a Tarefa dentro do Projeto
        rx_task = {"$regex": f"^{re.escape(task_name_norm)}", "$options": "i"}
        task_query = {"nome": rx_task, "projeto": DBRef(collection=COLL_PROJETOS, id=project_id)}
        task = mongo()[COLL_TAREFAS].find_one(task_query, {"_id": 1})
        
        if task:
            return str(task.get("_id"))
        else:
            print(f"[Helper] Tarefa '{task_name_norm}' não encontrada no projeto '{project_name_norm}'.")
            return None
    except Exception as e:
        print(f"Erro ao buscar task ID por nome: {e}")
        return None

# =========================
# Helpers de Download (PDF/XLSX)
# =========================
def fetch_bytes(url: str) -> bytes:
    if not url: raise ValueError("URL ausente")
    u = url.strip()
    if not u.lower().startswith(("http://", "https://")): raise ValueError("URL inválida")
    
    headers = {"User-Agent": GENERIC_USER_AGENT}
    
    is_google_export = "docs.google.com" in u and "export?format=xlsx" in u
    
    if is_google_export:
        with requests.get(u, timeout=TIMEOUT_S, allow_redirects=True, headers=headers, stream=True) as r:
            r.raise_for_status()
            ctype = (r.headers.get("Content-Type") or "").lower()
            if "openxmlformats-officedocument.spreadsheetml.sheet" not in ctype:
                raise ValueError(f"URL do Google não retornou um XLSX. Content-Type: {ctype}")
            content_bytes = io.BytesIO()
            for chunk in r.iter_content(chunk_size=8192):
                content_bytes.write(chunk)
            return content_bytes.getvalue()
    else:
        r = requests.get(u, timeout=TIMEOUT_S, allow_redirects=True, headers=headers)
        r.raise_for_status()
        return r.content

def _try_download_traced(url: str, timeout: int, user_agent: str):
    headers = {"User-Agent": user_agent, "Accept": "application/pdf, */*"}
    with requests.get(url, timeout=timeout, allow_redirects=True, headers=headers, stream=False) as r:
        status, ctype, final_url, content = r.status_code, (r.headers.get("Content-Type") or "").lower(), r.url, r.content or b""
        return {"status": status, "content_type": ctype, "final_url": final_url, "content": content, "is_pdf_signature": content[:4] == b"%PDF", "size": len(content)}

def normalize_sharepoint_pdf_url(u: str) -> str:
    try:
        pu = urlparse(u)
        if pu.netloc and "sharepoint.com" in pu.netloc and pu.path.endswith("/onedrive.aspx"):
            qs = parse_qs(pu.query or ""); idp = qs.get("id", [None])[0]
            if idp:
                raw_path = unquote(idp);
                if not raw_path.startswith("/"): raw_path = "/" + raw_path
                encoded_path = quote(raw_path, safe="/-_.()~")
                return f"{pu.scheme}://{pu.netloc}{encoded_path}"
    except Exception: pass
    return u

def normalize_gdrive_pdf_url(u: str) -> str:
    """Converte links de visualização do Google Drive em links de download direto."""
    try:
        m = re.search(r"/file/d/([a-zA-Z0-9_-]{28,})", u)
        if m:
            file_id = m.group(1)
            return f"https://drive.google.com/uc?export=download&id={file_id}"
        m_open = re.search(r"id=([a-zA-Z0-9_-]{28,})", u)
        if m_open and "drive.google.com" in u:
            file_id = m_open.group(1)
            return f"https://drive.google.com/uc?export=download&id={file_id}"
    except Exception:
        pass
    return u

def fetch_pdf_bytes(url: str):
    if not url: raise ValueError("URL ausente para PDF")
    u = url.strip()
    if not u.lower().startswith(("http://", "https://")): raise ValueError("URL inválida para PDF")
    u_norm = u
    if "sharepoint.com" in u:
        u_norm = normalize_sharepoint_pdf_url(u)
    elif "drive.google.com" in u:
        u_norm = normalize_gdrive_pdf_url(u)
    def _is_pdf(ctype: str, content: bytes) -> bool: 
        return ("pdf" in ctype.lower()) or ("octet-stream" in ctype.lower()) or (content[:4] == b"%PDF")
    variants = [("normalized", u_norm)]
    if u_norm != u:
        variants.append(("original", u))
    if "sharepoint.com" in u_norm.lower():
        if "download=1" not in u_norm.lower():
            sep = "&" if "?" in u_norm else "?"; variants.append(("download=1", u_norm + f"{sep}download=1"))
        pu = urlparse(u_norm); base = f"{pu.scheme}://{pu.netloc}"; src = quote(u_norm, safe=""); variants.append(("download.aspx", f"{base}/_layouts/15/download.aspx?SourceUrl={src}"))
    last = None
    for label, cand in variants:
        print(f"[DEBUG-Fetch] Tentando baixar (método: {label}): {cand[:100]}...")
        r = _try_download_traced(cand, TIMEOUT_S, PDF_USER_AGENT); last = r
        if "drive.google.com" in cand and "text/html" in r["content_type"].lower():
            try:
                print("[DEBUG-Fetch] GDrive retornou HTML. Verificando se é página de confirmação...")
                m_confirm = re.search(r'href="(/uc\?export=download&amp;confirm=[a-zA-Z0-9_&;-]+)"', r["content"].decode('utf-8', errors='ignore'))
                if m_confirm:
                    confirm_url = "https://drive.google.com" + m_confirm.group(1).replace("&amp;", "&")
                    print(f"[DEBUG-Fetch] GDrive precisa de confirmação. Tentando: {confirm_url[:100]}...")
                    r = _try_download_traced(confirm_url, TIMEOUT_S, PDF_USER_AGENT); last = r
            except Exception as e_parse:
                print(f"[DEBUG-Fetch] Falha ao parsear página de confirmação do GDrive: {e_parse}")
        if r["status"] == 200 and _is_pdf(r["content_type"], r["content"]): 
            print("[DEBUG-Fetch] Download concluído.")
            return r["content"]
    raise ValueError(f"Não foi possível obter PDF (último status={last['status'] if last else None}, content-type={last['content_type'] if last else None}).")

def clean_pdf_text(s: str) -> str:
    if not s: return s
    s = re.sub(r"([^\n])\n([^\n])", r"\1 \2", s)
    s = re.sub(r"[ \t\r\f\v]+", " ", s)
    s = re.sub(r"\s+([,;\.\!\?\:\)])", r"\1", s)
    s = re.sub(r"([,;\.\!\?\:])([^\s])", r"\1 \2", s)
    return s.strip()

def _anchor_regex_flex(label: str) -> re.Pattern:
    m = re.search(r"(?i)texto\.?(\d+)", label or "");
    if not m: return re.compile(r"(?!)")
    num = m.group(1); return re.compile(rf"(?i)\bTexto\.?{re.escape(num)}\.?\b[:\-]?\s*")

def extract_after_anchor_from_pdf(text_all: str, anchor_label: str, max_chars: int = 4000) -> str:
    if not text_all.strip(): return ""
    
    rx = _anchor_regex_flex(anchor_label); m = rx.search(text_all)
    if not m: return ""
    
    start = m.end()
    next_m = re.search(r"(?i)\bTexto\.?\d+\.?\b", text_all[start:])
    end = start + next_m.start() if next_m else len(text_all)
    
    return text_all[start:end].strip()[:max_chars].strip()

def xlsx_bytes_to_dataframe_preserving_hyperlinks(xlsx_bytes: bytes) -> pd.DataFrame:
    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=False); ws = wb.active
    headers: List[str] = [str(cell.value).strip() if cell.value is not None else "" for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=False))]
    hyperlink_map: Dict[str, str] = {}
    for hl in getattr(ws, "_hyperlinks", []) or []:
        try:
            ref = getattr(hl, "ref", None); target = getattr(hl, "target", None) or getattr(hl, "location", None)
            if not ref or not target: continue
            if ":" in ref:
                min_col, min_row, max_col, max_row = range_boundaries(ref)
                for r in range(min_row, max_row + 1):
                    for c in range(min_col, max_col + 1): hyperlink_map[ws.cell(row=r, column=c).coordinate] = target
            else: hyperlink_map[ref] = target
        except Exception: continue
    rows: List[Dict[str, Any]] = []
    col_idx_doc = headers.index("Documento Referência") if "Documento Referência" in headers else -1
    for row in ws.iter_rows(min_row=2, values_only=False):
        if all((c.value is None or str(c.value).strip() == "") for c in row): continue
        record: Dict[str, Any] = {}
        for i, cell in enumerate(row):
            header = headers[i] if i < len(headers) else f"col{i+1}"; val = cell.value
            if i == col_idx_doc:
                url = (getattr(cell.hyperlink, "target", None) if getattr(cell, "hyperlink", None) else None) or hyperlink_map.get(cell.coordinate)
                if not url and isinstance(val, str) and val.strip().lower().startswith(("http://", "https://")): url = val.strip()
                record[header] = url or (val if val is not None else "")
            else: record[header] = val if val is not None else ""
        rows.append(record)
    return pd.DataFrame(rows)

def extract_full_pdf_text(pdf_bytes: bytes) -> str:
    text_all = ""
    print("[DEBUG] Iniciando extract_full_pdf_text V13 (Super-Debug)...")
    try:
        print("[DEBUG] Tentando extração com PDFPLUMBER...")
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            text_all = "\n".join((page.extract_text(x_tolerance=3, y_tolerance=3) or "") for page in pdf.pages if page.extract_text())
        if text_all.strip(): 
            print("[DEBUG] PDFPLUMBER SUCESSO. Retornando texto.")
            print(f"[DEBUG] Amostra PDFPLUMBER: {text_all.strip()[:200]}")
            return text_all
        print("[DEBUG] PDFPLUMBER não retornou texto. Fallback para PyMuPDF/fitz.")
    except Exception as e:
        print(f"[DEBUG] PDFPLUMBER FALHOU com erro: {e}. Fallback para PyMuPDF/fitz.")
    try:
        print("[DEBUG] Tentando extração com FITZ (PyMuPDF)...")
        text_all_fitz = ""
        with fitz.open(stream=pdf_bytes, filetype="pdf") as doc:
            for page in doc:
                text_all_fitz += page.get_text("text") or ""
        if text_all_fitz.strip():
            print("[DEBUG] FITZ SUCESSO. Retornando texto.")
            print(f"[DEBUG] Amostra FITZ: {text_all_fitz.strip()[:200]}") 
            return text_all_fitz
        else:
            print("[DEBUG] FITZ não retornou texto. Retornando vazio.")
            return ""  
    except Exception as e_fitz:
         print(f"[DEBUG] FITZ FALHOU com erro: {e_fitz}")
         return ""
    
def extract_hidden_message(raw_text: str) -> str:
    if not raw_text: return ""
    secret_parts = []
    text = re.sub(r"([a-zA-Z])\s*\n\s*([a-zA-Z])", r"\1\2", raw_text)
    text = re.sub(r"\s+", " ", text)
    words = text.split()
    sentence_enders = (".", "!", "?", ":")
    for i, word in enumerate(words):
        clean_word = word.rstrip(f",;:.\"'{''.join(sentence_enders)}")
        if not clean_word or len(clean_word) < 1: continue
        if clean_word.islower(): continue
        if clean_word.isupper(): continue
        is_start_of_sentence = False
        if i == 0: is_start_of_sentence = True
        else:
            prev_word_raw = words[i-1]
            prev_word_clean = prev_word_raw.rstrip(f",;:.\"'{''.join(sentence_enders)}")
            if any(prev_word_raw.endswith(ender) for ender in sentence_enders): is_start_of_sentence = True
            elif not prev_word_clean.islower():
                if prev_word_clean.istitle() or prev_word_clean.isupper(): is_start_of_sentence = True
        if is_start_of_sentence and clean_word.istitle(): continue
        caps = re.sub(r"[^A-Z]", "", word) 
        if caps: secret_parts.append(caps)
    return " ".join(secret_parts)
    
# =========================
# Auth (Falar com API Render)
# =========================
_token_cache: Dict[str, Any] = {"access_token": None, "expires_at": 0, "user_id": None}
async def get_auth_header(client: httpx.AsyncClient) -> Dict[str, str]:
    if ACHEFLOW_MAIN_API_TOKEN:
        return {"Authorization": f"Bearer {ACHEFLOW_MAIN_API_TOKEN}"}
    now = time.time()
    if _token_cache.get("access_token") and now < _token_cache.get("expires_at", 0) - 30:
        return {"Authorization": f"Bearer {_token_cache['access_token']}"}
    if not TASKS_API_USERNAME or not TASKS_API_PASSWORD:
        raise HTTPException(status_code=401, detail="Nenhum token (ACHEFLOW_MAIN_API_TOKEN) ou credenciais (TASKS_API_USERNAME/PASSWORD) fornecidos para a API Principal.")
    token_url = urljoin(TASKS_API_BASE + "/", TASKS_API_TOKEN_PATH.lstrip("/"))
    data = {"username": TASKS_API_USERNAME, "password": TASKS_API_PASSWORD}
    headers = {"Content-Type": "application/x-www-form-urlencoded"}
    try:
        resp = await client.post(token_url, data=data, headers=headers, timeout=TIMEOUT_S)
        resp.raise_for_status()
        payload = resp.json()
        access_token = payload.get("access_token")
        expires_in  = int(payload.get("expires_in") or 3600)
        _token_cache["user_id"] = payload.get("id") or _token_cache.get("user_id")
        if not access_token: raise RuntimeError(f"Resposta de token sem access_token")
        _token_cache["access_token"] = access_token
        _token_cache["expires_at"] = time.time() + expires_in
        return {"Authorization": f"Bearer {access_token}"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Falha ao autenticar com a API Principal em {token_url}: {str(e)}")
async def get_api_auth_headers(client: httpx.AsyncClient, use_json: bool = True) -> Dict[str, str]:
    auth_header = await get_auth_header(client)
    if use_json:
        auth_header["Content-Type"] = "application/json"
    return auth_header

# =========================
# Lógica de Importação (do ai_api.py)
# =========================
class CreateTaskItem(BaseModel):
    titulo: str; descricao: Optional[str] = None; responsavel: Optional[str] = None
    deadline: Optional[str] = None; doc_ref: Optional[str] = None; prazo_data: Optional[str] = None

async def create_project_api(client: httpx.AsyncClient, data: Dict[str, Any]) -> Dict[str, Any]:
    url = f"{TASKS_API_BASE}{TASKS_API_PROJECTS_PATH}"
    auth_headers = await get_api_auth_headers(client, use_json=True)
    payload = pick(data, ["nome", "responsavel_id", "situacao", "prazo", "descricao", "categoria"])
    r = await client.post(url, json=payload, headers=auth_headers, timeout=TIMEOUT_S)
    if r.status_code not in (200, 201):
        raise httpx.HTTPStatusError(f"Erro da API do Render: {r.status_code}", request=r.request, response=r)
    return sanitize_doc(r.json())

async def create_task_api(client: httpx.AsyncClient, data: Dict[str, Any]) -> Dict[str, Any]:
    url = f"{TASKS_API_BASE}{TASKS_API_TASKS_PATH}"
    auth_headers = await get_api_auth_headers(client, use_json=True)
    payload = pick(data, ["nome", "projeto_id", "responsavel_id", "descricao", "prioridade", "status", "prazo", "documento_referencia", "concluido"])
    payload = {k: v for k, v in payload.items() if v is not None}
    r = await client.post(url, json=payload, headers=auth_headers, timeout=TIMEOUT_S)
    if r.status_code not in (200, 201):
        raise httpx.HTTPStatusError(f"Erro da API do Render: {r.status_code}", request=r.request, response=r)
    return sanitize_doc(r.json())

async def find_project_id_by_name(client: httpx.AsyncClient, projeto_nome: str) -> Optional[str]:
    url = f"{TASKS_API_BASE}{TASKS_API_PROJECTS_PATH}"
    auth_headers = await get_api_auth_headers(client, use_json=True)
    r = await client.get(url, headers=auth_headers, timeout=TIMEOUT_S)
    if r.status_code != 200: return None
    try:
        items = r.json()
        if isinstance(items, list):
            hit = next((p for p in items if str(p.get("nome")).lower() == projeto_nome.lower()), None)
            return (hit or {}).get("_id") if hit else None
    except Exception: return None
    return None

async def list_funcionarios(client: httpx.AsyncClient) -> List[Dict[str, Any]]:
    url = f"{TASKS_API_BASE}/funcionarios"
    auth_headers = await get_api_auth_headers(client, use_json=True)
    r = await client.get(url, headers=auth_headers, timeout=TIMEOUT_S)
    if r.status_code != 200: return []
    try: return r.json() if isinstance(r.json(), list) else []
    except Exception: return []

async def resolve_responsavel_id(
    client: httpx.AsyncClient, 
    nome_ou_email: Optional[str],
    default_user_id: Optional[str] = None
) -> Optional[str]:
    nome_ou_email = (nome_ou_email or "").strip()
    fallback_id = default_user_id if default_user_id else _token_cache.get("user_id")
    if nome_ou_email.lower() in ("eu", "mim", "me"):
        return default_user_id if default_user_id else _token_cache.get("user_id")
    if not nome_ou_email: 
        return fallback_id
    pessoas = await list_funcionarios(client)
    key = nome_ou_email.lower()
    if len(key) == 24 and all(c in '0123456789abcdef' for c in key):
        if any(p.get("_id") == key for p in pessoas): 
            return key
    perfect_match = None
    for p in pessoas:
        if str(p.get("email") or "").lower() == key:
            return p.get("_id")
        nome_clean = str(p.get('nome') or '').strip().lower()
        sobrenome_clean = str(p.get('sobrenome') or '').strip().lower()
        parts = [nome_clean, sobrenome_clean]
        full = " ".join(part for part in parts if part) 
        if full == key:
            perfect_match = p.get("_id")
    if perfect_match:
        return perfect_match
    for p in pessoas:
        nome_clean = str(p.get('nome') or '').strip().lower()
        sobrenome_clean = str(p.get('sobrenome') or '').strip().lower()
        if nome_clean == key:
            return p.get("_id")
        if sobrenome_clean == key:
            return p.get("_id")         
    return fallback_id

def duration_to_date(duracao: Optional[str]) -> str:
    base = datetime.utcnow().date()
    try:
        s = (duracao or "").strip().lower(); m = re.search(r"(\d+)", s)
        n = int(m.group(1)) if m else 7
    except Exception: n = 7
    return (base + timedelta(days=n)).isoformat()

def resolve_descricao_pdf(row) -> str:
    como, docrf = str(row.get("Como Fazer") or "").strip(), str(row.get("Documento Referência") or "").strip()
    if not como or not docrf or not re.search(r"(?i)\b((?:Doc\.?\s*)?(Texto\.?\d+))\b\.?", como): 
        return como
    try: 
        pdf_bytes = fetch_pdf_bytes(docrf)
        full_pdf_text = extract_full_pdf_text(pdf_bytes)
        if not full_pdf_text.strip():
            print(f"[DEBUG] PDF {docrf} não retornou texto. Usando '{como}' como fallback.")
            return como
    except Exception as e:
        print(f"[DEBUG] Falha ao baixar/processar PDF {docrf}: {e}. Usando '{como}' como fallback.")
        return como
    def _repl(m: re.Match) -> str:
        full_token, anchor = m.group(1), m.group(2)
        extracted = clean_pdf_text(extract_after_anchor_from_pdf(full_pdf_text, anchor))
        return extracted if extracted else full_token
    return re.sub(r"(?i)\b((?:Doc\.?\s*)?(Texto\.?\d+))\b\.?", _repl, como)

async def tasks_from_xlsx_logic(
    projeto_id: Optional[str],
    projeto_nome: Optional[str],
    user_id: Optional[str],
    create_project_flag: int,
    projeto_situacao: Optional[str],
    projeto_prazo: Optional[str],
    projeto_responsavel: Optional[str],
    projeto_descricao: Optional[str],
    projeto_categoria: Optional[str],
    xlsx_url: Optional[str],
    file_bytes: Optional[bytes]
) -> Dict[str, Any]:
    
    try:
        if file_bytes:
            df = xlsx_bytes_to_dataframe_preserving_hyperlinks(file_bytes)
        elif xlsx_url:
            xbytes = fetch_bytes(xlsx_url) 
            df = xlsx_bytes_to_dataframe_preserving_hyperlinks(xbytes)
        else:
            raise HTTPException(status_code=400, detail={"erro": "Nenhuma fonte de dados (file, xlsx_url, google_sheet_url) fornecida."})
    except Exception as e:
        raise HTTPException(status_code=422, detail={"erro": f"Falha ao processar planilha: {str(e)}"})

    required = {"Nome", "Como Fazer", "Documento Referência"}
    if not required.issubset(df.columns):
        missing = required - set(df.columns)
        raise HTTPException(status_code=400, detail={"erro": f"Colunas faltando: {', '.join(missing)}"})
    df["descricao_final"] = df.apply(resolve_descricao_pdf, axis=1)
    preview: List[Dict[str, Any]] = []
    latest_task_date: Optional[datetime.date] = None
    today_date = datetime.utcnow().date()
    for _, row in df.iterrows():
        prazo_col = str(row.get("Prazo") or "").strip()
        task_date_obj: Optional[datetime.date] = None
        if prazo_col:
            try:
                if re.match(r"^\d{2}/\d{2}/\d{4}$", prazo_col): d, m, y = prazo_col.split("/"); prazo_col = f"{y}-{m}-{d}"
                task_date_obj = datetime.strptime(prazo_col, "%Y-%m-%d").date()
            except Exception: prazo_col = None
        if not task_date_obj:
            duracao_txt = str(row.get("Duração") or "")
            try:
                s = (duracao_txt or "").strip().lower(); m = re.search(r"(\d+)", s)
                n = int(m.group(1)) if m else 7; task_date_obj = today_date + timedelta(days=n)
            except Exception: task_date_obj = today_date + timedelta(days=7)
            prazo_col = task_date_obj.isoformat()
        if latest_task_date is None or task_date_obj > latest_task_date:
            latest_task_date = task_date_obj
        preview.append({
            "titulo": str(row["Nome"]),
            "descricao": str(row.get("descricao_final") or ""),
            "responsavel": str(row.get("Responsavel") or ""), # <-- Isso será IGNORADO (propositalmente)
            "doc_ref": str(row.get("Documento Referência") or "").strip(),
            "prazo": prazo_col,
        })
    if not projeto_id and not projeto_nome:
        raise HTTPException(status_code=400, detail={"erro": "Para importar, forneça 'projeto_id' ou 'projeto_nome'."})
    async with httpx.AsyncClient() as client:
        proj_resp_id_unificado = await resolve_responsavel_id(client, projeto_responsavel, default_user_id=user_id)
        resolved_project_id: Optional[str] = projeto_id
        if not resolved_project_id and projeto_nome:
            resolved_project_id = await find_project_id_by_name(client, projeto_nome)
        if not resolved_project_id:
            if create_project_flag and projeto_nome:
                proj_prazo = (projeto_prazo or "").strip()
                if not proj_prazo:
                    proj_prazo = (latest_task_date or (today_date + timedelta(days=30))).isoformat()
                proj = await create_project_api(client, {
                    "nome": projeto_nome, 
                    "responsavel_id": proj_resp_id_unificado,
                    "situacao": (projeto_situacao or "Em planejamento").strip(),
                    "prazo": proj_prazo, "descricao": projeto_descricao, "categoria": projeto_categoria
                })
                resolved_project_id = proj.get("_id") or proj.get("id")
            else:
                raise HTTPException(status_code=404, detail={"erro": f"Projeto '{projeto_nome}' não encontrado. Para criar, envie 'create_project_flag=1'."})
        created, errors = [], []
        for item in preview:            
            try:
                created.append(await create_task_api(client, {
                    "nome": item["titulo"], "descricao": item["descricao"],
                    "projeto_id": resolved_project_id, 
                    "responsavel_id": proj_resp_id_unificado, # <-- MUDANÇA 4: Usar ID unificado
                    "prazo": item["prazo"],
                    "documento_referencia": item["doc_ref"],
                    "status": "não iniciada", "prioridade": "média"
                }))
            except Exception as e:
                errors.append({"erro": str(e), "titulo": item["titulo"]})
                
    return {"mode": "assigned", "projeto_id": resolved_project_id, "criados": created, "total": len(created), "erros": errors}

# =========================
# Lógica da IA (do vertex_ai_service.py)
# =========================

SYSTEM_PROMPT = """
====================================================================
REGRA DE OURO: AÇÃO IMEDIATA (NÃO FALE, FAÇA)
====================================================================
Sua missão é executar ferramentas. Esta é sua prioridade número 1.
VOCÊ ESTÁ ESTRITAMENTE PROIBIDO DE RESPONDER "Sim, claro", "Ok, atualizando..." ou "Pronto!" ANTES de chamar uma ferramenta.

**NOVA REGRA: TRATAMENTO DE LINKS (HTTP/HTTPS)**
Se a mensagem do usuário contiver uma URL (http:// ou https://) E uma pergunta sobre ela (como 'o que diz este pdf?', 'resuma este link', 'qual o enigma deste arquivo?', 'importe esta planilha'), sua prioridade MÁXIMA é usar uma das ferramentas de URL:
- `get_pdf_content_from_url` (para ler PDFs)
- `solve_pdf_enigma_from_url` (para enigmas em PDFs)
- `get_pdf_content_from_xlsx_url` (para ler PDFs DENTRO de planilhas)
- `solve_enigma_from_xlsx_url` (para enigmas DENTRO de planilhas)
- `import_project_from_url` (para importar de XLSX/Google Sheets)
NÃO RESPONDA que não pode acessar links. As ferramentas fazem isso por você. Chame a ferramenta apropriada IMEDIATAMENTE.

**ERRADO (NUNCA FAÇA ISSO, ISSO É UMA FALHA GRAVE):**
Usuário: "altere a tarefa X para 'concluída'"
IA: "Claro! Alterando a tarefa X para 'concluída'."

**ERRADO TAMBÉM (FALHA GRAVE):**
Usuário: "O que diz este PDF? https://drive.google.com/..."
IA: "Desculpe, eu não consigo acessar links externos." (ERRADO! Você devia ter chamado `get_pdf_content_from_url`)

**CORRETO (FAÇA ISSO):**
Usuário: "altere a tarefa X para 'concluída'"
IA: (Imediatamente chama a ferramenta `update_task(task_name="X", patch={{"status": "concluída"}})`)
IA: (APÓS a ferramenta retornar 'ok') "Pronto! A tarefa X foi atualizada."

**CORRETO TAMBÉM (FAÇA ISSO):**
Usuário: "O que diz este PDF? https://drive.google.com/..."
IA: (Imediatamente chama a ferramenta `get_pdf_content_from_url(url="https://drive.google.com/...")`)
IA: (APÓS a ferramenta retornar o texto) "O PDF fala sobre..."

Somente após a execução da ferramenta você pode dar uma resposta amigável.
====================================================================
Você é o "Ache", um assistente de produtividade virtual da plataforma Ache Flow.
Sua missão é ajudar colaboradores(as) como {nome_usuario} (email: {email_usuario}, id: {id_usuario}) a entender e gerenciar tarefas, projetos e prazos.
====================================================================
REGRAS DE COLETA DE DADOS (PARA CRIAR/EDITAR)
====================================================================

**REGRA DE DESAMBIGUAÇÃO (A MAIS IMPORTANTE):**
- Se o usuário pedir para 'criar um projeto' E TAMBÉM fornecer uma URL (.xlsx ou Google Sheets) na *mesma* mensagem, sua ÚNICA AÇÃO deve ser chamar a ferramenta `import_project_from_url`.
- Ignore a ferramenta `create_project` nesse caso.
- Esta é a regra prioritária.

Sua tarefa é preencher os argumentos para as ferramentas.
**REGRA PRINCIPAL:** Sempre tente extrair os parâmetros (como nome, prazo, etc.) da ÚLTIMA MENSAGEM DO USUÁRIO.
- **SE** você conseguir extrair TODOS os argumentos **OBRIGATÓRIOS** (como `nome`, `prazo`, `situacao`, `responsavel`):
    - **NÃO PERGUNTE NADA MAIS.** Chame a ferramenta imediatamente.
    - Use `None` (ou simplesmente omita) para quaisquer argumentos **OPCIONAIS** (como `projeto_descricao` ou `projeto_categoria`) que não foram fornecidos.
- **SE** algum argumento **OBRIGATÓÓRIO** estiver faltando:
    - **AÍ SIM,** pergunte APENAS pelos argumentos **OBRIGATÓRIOS** que faltam.
    - **NÃO** pergunte por argumentos opcionais.
**REGRA DE AÇÃO DIRETA (A MAIS IMPORTANTE):**
- **NUNCA** responda ao usuário com uma "confirmação" antes de agir.
- **ERRADO (NÃO FAÇA ISSO):** O usuário diz "prazo 31-12-2025". Você responde: "OK. Criando projeto com prazo 31-12-2025."
- **CORRETO (FAÇA ISSO):** O usuário diz "prazo 31-12-2025". Você *imediatamente* chama a ferramenta `create_project(...)` em segundo plano. Somente *depois* que a ferramenta retornar `{{"ok": True, "data": ...}}`, você responde ao usuário: "Projeto criado com sucesso! 🙂"
- Se o usuário disser "isso" ou "sim" para confirmar, isso é sua instrução para **CHAMAR A FERRAMENTA**, não para falar mais.
**1. PARA: `create_project` (Criar Projeto ÚNICO):**
* **Argumentos OBRIGATÓRIOS:** `nome`, `situacao`, `prazo` (DD-MM-AAAA), `responsavel` (nome ou email).
* **Argumentos Opcionais:** `descricao`, `categoria`.
**2. PARA: `import_project_from_url` (Importar Projeto):**
* **Argumentos OBRIGATÓRIOS:** `xlsx_url`, `projeto_nome`, `projeto_situacao`, `projeto_prazo`, `projeto_responsavel`.
* **Argumentos Opcionais:** `projeto_descricao`, `projeto_categoria`.
**3. PARA: `update_project` (Atualizar Projeto):**
* **Se faltar:** O `patch` (o que mudar). O nome ou ID do projeto geralmente já é conhecido.
* **Exemplo:** Se o usuário disser "vamos alterar o projeto Pega-Pega", você DEVE perguntar: "Claro! O que você gostaria de mudar no projeto 'Pega-Pega' (nome, situação, prazo, etc.)?"
* **NÃO** pergunte pelo ID se o nome já foi dado. A ferramenta encontrará pelo nome.
(O resto das regras de update_project, update_task e DADOS DE CONTEXTO permanecem iguais)
====================================================================
REGRAS DE RESPOSTA (AGORA SECUNDÁRIAS)
====================================================================
**REGRA DE OURO: NÃO INVENTE DADOS.**
- Se uma ferramenta for usada e retornar uma lista vazia (como `[]`), um valor 0, "não encontrado", ou um erro, sua resposta DEVE ser "Não encontrei [o que foi pedido]" ou "Desculpe, não consegui processar isso. [Mensagem de Erro]".
- NUNCA, SOB NENHUMA CIRCUNSTÂNCIA, invente nomes de projetos, tarefas, pessoas ou frases secretas.
**REGRA ANTI-CÓDIGO: VOCÊ É UM ASSISTENTE, NÃO UM PROGRAMADOR.**
- Sua resposta para o usuário NUNCA deve ser um trecho de código (`print()`, JSON, etc).
- Sua tarefa é: 1º *chamar* a ferramenta, 2º *esperar* o resultado, e 3º *depois* formular uma resposta em português.
- Se você responder com `print(defaultapi.createproject...)`, você falhou gravemente.
- Você NUNCA deve inventar prefixos como `defaultapi` ou `print()`.
1.  **REGRA DE FERRAMENTAS (PRIORIDADE 1):** Sua prioridade MÁXIMA é usar ferramentas.
    * **REFORÇO CRÍTICO:** Ao 'criar', 'atualizar' ou 'importar', você está PROIBIDO de responder "Projeto criado" ou "Tarefa atualizada" sem ANTES chamar a ferramenta e receber a confirmação.
    * **REGRA DE IMPORTAÇÃO (DESAMBIGUAÇÃO):** <-- Esta regra foi movida para o topo da seção "COLETA DE DADOS".
    * NUNCA pergunte "Posso buscar?". Apenas execute a ferramenta e retorne a resposta.
    * Sempre que usar uma ferramenta, resuma o resultado em português claro. NUNCA mostre nomes de funções (como 'list_all_projects') ou código.
2.  **REGRA DE CONHECIMENTO GERAL (PRIORIDADE 2):** Se a pergunta NÃO PUDER ser respondida por NENHUMA ferramenta, use seu conhecimento pré-treinado.
    * Você NÃO precisa de acesso à internet para isso. Responda diretamente.
3.  **REGRA DE AMBIGUIDADE:** Se uma pergunta for ambígua (ex: "o que é um diferencial?"), responda com seu conhecimento geral.
4.  **REGRA DE FORMATAÇÃO:**
    * Fale sempre em português (PT-BR), de forma simpática.
    * NUNCA use markdown, asteriscos (*), negrito, ou blocos de código.
    * Use hífens simples para listas.
====================================================================
DADOS DE CONTEXTO
====================================================================
-   **Usuário Atual:** {nome_usuario} (ID: {id_usuario})
-   **Interpretação de "Eu":** Se o usuário disser "eu", "para mim", "sou eu", **E NENHUM OUTRO NOME FOR DADO**, use a palavra "eu" no campo 'responsavel'. Se um nome explícito (ex: "Lucas Rodrigues") for fornecido, ele tem prioridade.
-   **Datas:** Hoje é {data_hoje}. "Este mês" vai de {inicio_mes} até {fim_mes}.
-   **Formato de Data:** Sempre que pedir uma data, peça em **DD-MM-AAAA**. Você deve converter internamente para **AAAA-MM-DD** antes de usar nas ferramentas.
"""

def list_all_projects(top_k: int = 500) -> List[Dict[str, Any]]:
    employee_map = _get_employee_map()
    projects_raw = mongo()[COLL_PROJETOS].find({}).sort("prazo", 1).limit(top_k)
    projects_clean = [sanitize_doc(p) for p in projects_raw]
    return [_enrich_doc_with_responsavel(p, employee_map) for p in projects_clean]

def list_all_tasks(top_k: int = 2000) -> List[Dict[str, Any]]:
    employee_map = _get_employee_map()
    tasks_raw = mongo()[COLL_TAREFAS].find({}).sort("prazo", 1).limit(top_k)
    tasks_clean = [sanitize_doc(t) for t in tasks_raw]
    return [_enrich_doc_with_responsavel(t, employee_map) for t in tasks_clean]

def list_all_funcionarios(top_k: int = 500) -> List[Dict[str, Any]]:
    return [sanitize_doc(x) for x in mongo()[COLL_FUNCIONARIOS].find({}).sort("nome", 1).limit(top_k)]

def list_tasks_by_deadline_range(start: str, end: str, top_k: int = 50) -> List[Dict[str, Any]]:
    employee_map = _get_employee_map()
    tasks_raw = mongo()[COLL_TAREFAS].find({"prazo": {"$gte": start, "$lte": end}}).sort("prazo", 1).limit(top_k)
    tasks_clean = [sanitize_doc(t) for t in tasks_raw]
    return [_enrich_doc_with_responsavel(t, employee_map) for t in tasks_clean]

def list_projects_by_status(status: str, top_k: int = 50) -> List[Dict[str, Any]]:
    status_norm = (status or "").strip()
    if not status_norm: return []
    rx = {"$regex": f"^{re.escape(status_norm)}$", "$options": "i"}
    employee_map = _get_employee_map()
    projects_raw = mongo()[COLL_PROJETOS].find({"situacao": rx}).sort("prazo", 1).limit(top_k)
    projects_clean = [sanitize_doc(p) for p in projects_raw]
    return [_enrich_doc_with_responsavel(p, employee_map) for p in projects_clean]

def upcoming_deadlines(days: int = 14, top_k: int = 50) -> List[Dict[str, Any]]:
    today_iso = iso_date(today()); limit_date = (today() + timedelta(days=days)).date().isoformat()
    employee_map = _get_employee_map()
    tasks_raw = mongo()[COLL_TAREFAS].find({"prazo": {"$gte": today_iso, "$lte": limit_date}}).sort("prazo", 1).limit(top_k)
    tasks_clean = [sanitize_doc(t) for t in tasks_raw]
    return [_enrich_doc_with_responsavel(t, employee_map) for t in tasks_clean]

def count_all_projects() -> int:
    try:
        return mongo()[COLL_PROJETOS].count_documents({})
    except Exception as e:
        print(f"Erro ao contar projetos: {e}")
        return -1

def count_all_tasks() -> int:
    try:
        return mongo()[COLL_TAREFAS].count_documents({})
    except Exception as e:
        print(f"Erro ao contar tarefas: {e}")
        return -1

def find_employee_with_most_tasks() -> Dict[str, Any]:
    """
    Usa um pipeline de agregação do MongoDB para encontrar o funcionário
    com o maior número de tarefas atribuídas. (Versão 2 - Robusta)
    """
    try:
        pipeline = [
            {"$match": {"responsavel": {"$exists": True, "$ne": None, "$ne": ""}}},
            
            {"$group": {"_id": "$responsavel", "count": {"$sum": 1}}},
            
            {"$sort": {"count": -1}},
            
            {"$limit": 1},
            
            {
                "$lookup": {
                    "from": COLL_FUNCIONARIOS,
                    "localField": "_id",
                    "foreignField": "_id",
                    "as": "funcionario_doc"
                }
            },
            
            {"$unwind": {"path": "$funcionario_doc", "preserveNullAndEmptyArrays": True}},
            
            {
                "$project": {
                    "_id": 0,
                    "total_tarefas": "$count",
                    "nome_funcionario": {
                        "$trim": {
                            "input": {
                                "$concat": [
                                    {"$ifNull": ["$funcionario_doc.nome", "(Funcionário não encontrado)"]},
                                    " ",
                                    {"$ifNull": ["$funcionario_doc.sobrenome", ""]}
                                ]
                            }
                        }
                    }
                }
            }
        ]
        
        result = list(mongo()[COLL_TAREFAS].aggregate(pipeline))
        
        if not result:
            return {"erro": "Não foi possível encontrar funcionários com tarefas (ou nenhuma tarefa está atribuída)."}
        return result[0]

    except Exception as e:
        print(f"Erro ao buscar funcionário com mais tarefas: {e}")
        return {"erro": str(e)}

def _find_employee_id_by_name_sync(name_or_email: str) -> Optional[ObjectId]:
    """
    Helper SÍNCRONO para encontrar o ID de um funcionário no MongoDB
    pelo nome, sobrenome, nome completo ou email. (Versão 2 - Corrigida)
    """
    key = (name_or_email or "").strip().lower()
    if not key: return None
    
    try:
        # 1. Tentar por email (sem-mudança)
        query_email = {"email": {"$regex": f"^{re.escape(key)}$", "$options": "i"}}
        emp = mongo()[COLL_FUNCIONARIOS].find_one(query_email, {"_id": 1})
        if emp: return emp.get("_id")

        # 2. Tentar por ID (sem-mudança)
        if len(key) == 24 and all(c in '0123456789abcdef' for c in key):
            emp = mongo()[COLL_FUNCIONARIOS].find_one({"_id": to_oid(key)}, {"_id": 1})
            if emp: return emp.get("_id")

        # 3. Tentar por nome completo (LÓGICA CORRIGIDA)
        parts = key.split()
        if len(parts) > 1:
            # --- INÍCIO DA CORREÇÃO ---
            
            # Tentativa 1 (Mais Comum): "Nome Composto" "Sobrenome"
            # Ex: "Ana Luiza" (nome) "Dourado" (sobrenome)
            nome_parts_1 = parts[:-1]  # Tudo exceto o último
            nome_str_1 = " ".join(nome_parts_1)
            sobrenome_str_1 = parts[-1] # Apenas o último
            
            nome_rx_1 = {"$regex": f"^{re.escape(nome_str_1)}$", "$options": "i"}
            sobrenome_rx_1 = {"$regex": f"^{re.escape(sobrenome_str_1)}$", "$options": "i"}

            query_full_1 = {"nome": nome_rx_1, "sobrenome": sobrenome_rx_1}
            emp_1 = mongo()[COLL_FUNCIONARIOS].find_one(query_full_1, {"_id": 1})
            if emp_1: return emp_1.get("_id")
            
            # Tentativa 2 (Fallback): "Nome" "Sobrenome Composto"
            # Ex: "Ana" (nome) "Luiza Dourado" (sobrenome)
            if len(parts) > 2: # Só tenta se houver nome do meio
                nome_str_2 = parts[0] # Apenas o primeiro
                sobrenome_str_2 = " ".join(parts[1:]) # Todo o resto
                
                nome_rx_2 = {"$regex": f"^{re.escape(nome_str_2)}$", "$options": "i"}
                sobrenome_rx_2 = {"$regex": f"^{re.escape(sobrenome_str_2)}$", "$options": "i"}
                
                query_full_2 = {"nome": nome_rx_2, "sobrenome": sobrenome_rx_2}
                emp_2 = mongo()[COLL_FUNCIONARIOS].find_one(query_full_2, {"_id": 1})
                if emp_2: return emp_2.get("_id")
            # --- FIM DA CORREÇÃO ---

        # 4. Tentar por primeiro nome (APENAS se for um resultado único)
        first_name_rx = {"$regex": f"^{re.escape(parts[0])}$", "$options": "i"}
        query_first = {"nome": first_name_rx}
        emp_cursor = mongo()[COLL_FUNCIONARIOS].find(query_first, {"_id": 1})
        emps = list(emp_cursor)
        if len(emps) == 1: # Só retorna se for uma correspondência inequívoca
            return emps[0].get("_id")

        print(f"[_find_employee_id_by_name_sync] Não foi possível encontrar funcionário por '{key}'")
        return None
    except Exception as e:
        print(f"Erro ao buscar ID de funcionário por nome: {e}")
        return None
    
def _get_employee_doc_by_name(name_or_email: str) -> Optional[Dict[str, Any]]:
    """
    Helper síncrono para buscar o DOCUMENTO COMPLETO de um funcionário.
    Reutiliza a lógica de busca por ID.
    """
    emp_id = _find_employee_id_by_name_sync(name_or_email)
    if not emp_id:
        return None
    try:
        # Retorna o documento completo (sem 'sanitize_doc' ainda)
        return mongo()[COLL_FUNCIONARIOS].find_one({"_id": emp_id})
    except Exception as e:
        print(f"Erro ao buscar documento completo do funcionário: {e}")
        return None

def get_employee_email(employee_name: str) -> Dict[str, Any]:
    """
    Encontra o email de um funcionário específico.
    """
    emp_doc = _get_employee_doc_by_name(employee_name)
    if not emp_doc:
        return {"erro": f"Funcionário '{employee_name}' não encontrado."}
    
    return {
        "nome": f"{emp_doc.get('nome', '')} {emp_doc.get('sobrenome', '')}".strip(),
        "email": emp_doc.get("email", "Email não informado")
    }

def get_employee_position(employee_name: str) -> Dict[str, Any]:
    """
    Encontra o cargo de um funcionário específico.
    (Assume que o campo se chama 'cargo')
    """
    emp_doc = _get_employee_doc_by_name(employee_name)
    if not emp_doc:
        return {"erro": f"Funcionário '{employee_name}' não encontrado."}
    
    return {
        "nome": f"{emp_doc.get('nome', '')} {emp_doc.get('sobrenome', '')}".strip(),
        "cargo": emp_doc.get("cargo", "Cargo não informado")
    }

def get_employee_department(employee_name: str) -> Dict[str, Any]:
    """
    Encontra o departamento de um funcionário específico.
    (Assume que o campo se chama 'departamento')
    """
    emp_doc = _get_employee_doc_by_name(employee_name)
    if not emp_doc:
        return {"erro": f"Funcionário '{employee_name}' não encontrado."}
    
    return {
        "nome": f"{emp_doc.get('nome', '')} {emp_doc.get('sobrenome', '')}".strip(),
        "departamento": emp_doc.get("departamento", "Departamento não informado")
    }

def get_employee_tenure(employee_name: str) -> Dict[str, Any]:
    """
    Calcula o tempo de casa (preciso) e retorna a data de admissão.
    (Versão V3: Unificada + Cálculo de dias)
    """
    emp_doc = _get_employee_doc_by_name(employee_name)
    if not emp_doc:
        return {"erro": f"Funcionário '{employee_name}' não encontrado."}

    nome_fmt = f"{emp_doc.get('nome', '')} {emp_doc.get('sobrenome', '')}".strip()
    
    data_admissao_val = emp_doc.get("dataCadastro") 
    
    if not data_admissao_val:
        return {"nome": nome_fmt, "tempo_de_casa": "Data de cadastro não informada"}

    try:
        # Converte para 'date' (ignora hora/fuso)
        if isinstance(data_admissao_val, datetime):
            start_date = data_admissao_val.date()
        else:
            start_date = datetime.fromisoformat(str(data_admissao_val).split("T")[0]).date()
        
        today = datetime.utcnow().date()
        delta = today - start_date
        
        if delta.days < 0:
            return {
                "nome": nome_fmt, 
                "tempo_de_casa": f"ainda não começou",
                "data_admissao": start_date.strftime('%d/%m/%Y'),
                "data_admissao_iso": start_date.isoformat()
            }
        
        if delta.days == 0:
            return {
                "nome": nome_fmt, 
                "tempo_de_casa": "começou hoje",
                "data_admissao": start_date.strftime('%d/%m/%Y'),
                "data_admissao_iso": start_date.isoformat()
            }

        anos = delta.days // 365
        dias_restantes_ano = delta.days % 365
        meses = dias_restantes_ano // 30  # Aproximação, mas consistente
        dias = dias_restantes_ano % 30

        parts = []
        if anos > 1: parts.append(f"{anos} anos")
        elif anos == 1: parts.append("1 ano")

        if meses > 1: parts.append(f"{meses} meses")
        elif meses == 1: parts.append("1 mês")

        if dias > 1: parts.append(f"{dias} dias")
        elif dias == 1: parts.append("1 dia")

        tenure_str = ""
        if len(parts) == 0:
            tenure_str = "começou hoje"
        elif len(parts) == 1:
            tenure_str = parts[0]
        elif len(parts) == 2:
            tenure_str = f"{parts[0]} e {parts[1]}"
        else: # 3 partes (anos, meses, dias)
            tenure_str = f"{parts[0]}, {parts[1]} e {parts[2]}"
        
        return {
            "nome": nome_fmt, 
            "tempo_de_casa": tenure_str,
            "data_admissao": start_date.strftime('%d/%m/%Y'), # Formato da imagem 2
            "data_admissao_iso": start_date.isoformat()
        }
    
    except Exception as e:
        print(f"Erro ao calcular tempo de casa para {nome_fmt}: {e}")
        return {"nome": nome_fmt, "tempo_de_casa": f"Data de cadastro inválida ({data_admissao_val})"}
        
def find_employee_with_least_tasks() -> Dict[str, Any]:
    """
    Encontra o funcionário com o MENOR número de tarefas (mínimo 1).
    """
    try:
        pipeline = [
            {"$match": {"responsavel": {"$exists": True, "$ne": None, "$ne": ""}}},
            {"$group": {"_id": "$responsavel", "count": {"$sum": 1}}},
            # A ÚNICA MUDANÇA É AQUI: 1 (ascendente)
            {"$sort": {"count": 1}},
            {"$limit": 1},
            {"$lookup": {"from": COLL_FUNCIONARIOS, "localField": "_id", "foreignField": "_id", "as": "funcionario_doc"}},
            {"$unwind": {"path": "$funcionario_doc", "preserveNullAndEmptyArrays": True}},
            {"$project": {
                "_id": 0,
                "total_tarefas": "$count",
                "nome_funcionario": {
                    "$trim": {
                        "input": {
                            "$concat": [
                                {"$ifNull": ["$funcionario_doc.nome", "(Funcionário não encontrado)"]},
                                " ",
                                {"$ifNull": ["$funcionario_doc.sobrenome", ""]}
                            ]
                        }
                    }
                }
            }}
        ]
        result = list(mongo()[COLL_TAREFAS].aggregate(pipeline))
        if not result:
            return {"erro": "Não foi possível encontrar funcionários com tarefas (ou nenhuma tarefa está atribuída)."}
        return result[0]
    except Exception as e:
        print(f"Erro ao buscar funcionário com menos tarefas: {e}")
        return {"erro": str(e)}

def find_employee_with_most_projects() -> Dict[str, Any]:
    """
    Encontra o funcionário com o MAIOR número de projetos.
    """
    try:
        pipeline = [
            {"$match": {"responsavel": {"$exists": True, "$ne": None, "$ne": ""}}},
            {"$group": {"_id": "$responsavel", "count": {"$sum": 1}}},
            {"$sort": {"count": -1}}, # -1 = Mais
            {"$limit": 1},
            {"$lookup": {"from": COLL_FUNCIONARIOS, "localField": "_id", "foreignField": "_id", "as": "funcionario_doc"}},
            {"$unwind": {"path": "$funcionario_doc", "preserveNullAndEmptyArrays": True}},
            {"$project": {
                "_id": 0,
                "total_projetos": "$count", # <- Mudei o nome da chave
                "nome_funcionario": {
                    "$trim": {
                        "input": {
                            "$concat": [
                                {"$ifNull": ["$funcionario_doc.nome", "(Funcionário não encontrado)"]},
                                " ",
                                {"$ifNull": ["$funcionario_doc.sobrenome", ""]}
                            ]
                        }
                    }
                }
            }}
        ]
        # MUDANÇA AQUI: Consulta COLL_PROJETOS
        result = list(mongo()[COLL_PROJETOS].aggregate(pipeline))
        if not result:
            return {"erro": "Não foi possível encontrar funcionários com projetos (ou nenhum projeto está atribuído)."}
        return result[0]
    except Exception as e:
        print(f"Erro ao buscar funcionário com mais projetos: {e}")
        return {"erro": str(e)}

def find_employee_with_least_projects() -> Dict[str, Any]:
    """
    Encontra o funcionário com o MENOR número de projetos (mínimo 1).
    """
    try:
        pipeline = [
            {"$match": {"responsavel": {"$exists": True, "$ne": None, "$ne": ""}}},
            {"$group": {"_id": "$responsavel", "count": {"$sum": 1}}},
            {"$sort": {"count": 1}}, # 1 = Menos
            {"$limit": 1},
            {"$lookup": {"from": COLL_FUNCIONARIOS, "localField": "_id", "foreignField": "_id", "as": "funcionario_doc"}},
            {"$unwind": {"path": "$funcionario_doc", "preserveNullAndEmptyArrays": True}},
            {"$project": {
                "_id": 0,
                "total_projetos": "$count", # <- Mudei o nome da chave
                "nome_funcionario": {
                    "$trim": {
                        "input": {
                            "$concat": [
                                {"$ifNull": ["$funcionario_doc.nome", "(Funcionário não encontrado)"]},
                                " ",
                                {"$ifNull": ["$funcionario_doc.sobrenome", ""]}
                            ]
                        }
                    }
                }
            }}
        ]
        # MUDANÇA AQUI: Consulta COLL_PROJETOS
        result = list(mongo()[COLL_PROJETOS].aggregate(pipeline))
        if not result:
            return {"erro": "Não foi possível encontrar funcionários com projetos (ou nenhum projeto está atribuído)."}
        return result[0]
    except Exception as e:
        print(f"Erro ao buscar funcionário com menos projetos: {e}")
        return {"erro": str(e)}

def count_tasks_by_employee_name(employee_name: str) -> Dict[str, Any]:
    """
    Conta o número de tarefas de um funcionário específico, buscando-o pelo nome.
    """
    emp_id = _find_employee_id_by_name_sync(employee_name)
    if not emp_id:
        return {"erro": f"Funcionário '{employee_name}' não encontrado."}
    
    try:
        # Pega o nome formatado (para o caso de "eu" ser usado)
        emp_doc = mongo()[COLL_FUNCIONARIOS].find_one({"_id": emp_id}, {"nome": 1, "sobrenome": 1})
        nome_fmt = f"{emp_doc.get('nome', '')} {emp_doc.get('sobrenome', '')}".strip()
        
        count = mongo()[COLL_TAREFAS].count_documents(
            {"responsavel": DBRef(collection=COLL_FUNCIONARIOS, id=emp_id)}
        )
        return {"nome_funcionario": nome_fmt, "total_tarefas": count}
    except Exception as e:
        return {"erro": str(e)}

def count_projects_by_employee_name(employee_name: str) -> Dict[str, Any]:
    """
    Conta o número de projetos de um funcionário específico, buscando-o pelo nome.
    """
    emp_id = _find_employee_id_by_name_sync(employee_name)
    if not emp_id:
        return {"erro": f"Funcionário '{employee_name}' não encontrado."}
    
    try:
        # Pega o nome formatado
        emp_doc = mongo()[COLL_FUNCIONARIOS].find_one({"_id": emp_id}, {"nome": 1, "sobrenome": 1})
        nome_fmt = f"{emp_doc.get('nome', '')} {emp_doc.get('sobrenome', '')}".strip()
        
        count = mongo()[COLL_PROJETOS].count_documents(
            {"responsavel": DBRef(collection=COLL_FUNCIONARIOS, id=emp_id)}
        )
        return {"nome_funcionario": nome_fmt, "total_projetos": count}
    except Exception as e:
        return {"erro": str(e)}

def count_projects_by_status(status: str) -> int:
    status_norm = (status or "").strip()
    if not status_norm: return 0
    try:
        rx = {"$regex": f"^{re.escape(status_norm)}$", "$options": "i"}
        return mongo()[COLL_PROJETOS].count_documents({"situacao": rx})
    except Exception as e:
        print(f"Erro ao contar projetos por status: {e}")
        return -1
    
async def update_project(
    patch: Dict[str, Any],
    project_id: Optional[str] = None, 
    project_name: Optional[str] = None,
    user_id: Optional[str] = None
) -> Dict[str, Any]:
    async with httpx.AsyncClient() as client:
        resolved_pid = project_id
        if not resolved_pid and project_name:
            resolved_pid = await find_project_id_by_name(client, project_name)
        if not resolved_pid:
            raise ValueError(f"Projeto '{project_name or project_id}' não encontrado ou ID/Nome não fornecido.")
        auth_headers = await get_api_auth_headers(client, use_json=True)
        allowed = {"nome", "descricao", "categoria", "situacao"}
        payload = {k: v for k, v in patch.items() if k in allowed and v is not None}
        if "prazo" in patch and patch["prazo"]:
            payload["prazo"] = _parse_date_robust(patch["prazo"])
        if "responsavel" in patch and patch["responsavel"]:
            resp_id = await resolve_responsavel_id(client, patch["responsavel"], default_user_id=user_id)
            payload["responsavel_id"] = resp_id
        if not payload: raise ValueError("Nenhum campo válido para atualizar ('patch' vazio).")
        url = f"{TASKS_API_BASE}/projetos/{resolved_pid}" 
        resp = await client.put(url, json=payload, headers=auth_headers)
        resp.raise_for_status(); return sanitize_doc(resp.json())

async def create_project(doc: Dict[str, Any], user_id: Optional[str] = None) -> Dict[str, Any]:
    async with httpx.AsyncClient() as client:
        data = pick(doc, ["nome", "situacao", "descricao", "categoria"])
        if not data.get("nome"): raise ValueError("nome é obrigatório")
        prazo_raw = doc.get("prazo")
        if prazo_raw:
            data["prazo"] = _parse_date_robust(prazo_raw)
        responsavel_str = doc.get("responsavel") 
        resolved_id = await resolve_responsavel_id(client, responsavel_str, default_user_id=user_id)
        data["responsavel_id"] = resolved_id
        return await create_project_api(client, data)
          
async def create_task(doc: Dict[str, Any]) -> Dict[str, Any]:
    async with httpx.AsyncClient() as client:
        data = pick(doc, ["nome", "projeto_id", "responsavel_id", "descricao", "prioridade", "status", "prazo", "documento_referencia", "concluido"])
        if not data.get("nome"): raise ValueError("nome é obrigatório")
        return await create_task_api(client, data)

async def update_task(
    patch: Dict[str, Any],
    task_id: Optional[str] = None, 
    task_name: Optional[str] = None,
    project_name: Optional[str] = None,
    user_id: Optional[str] = None
) -> Dict[str, Any]:
    
    async with httpx.AsyncClient() as client:
        resolved_tid = task_id
        
        if not resolved_tid and task_name and project_name:
            print(f"[update_task] ID não fornecido. Buscando por nome: '{task_name}' em '{project_name}'...")
            resolved_tid = _find_task_id_by_name_sync(task_name, project_name)
        
        if not resolved_tid:
            raise ValueError(f"Tarefa '{task_name or task_id}' não encontrada. Forneça um ID de tarefa válido ou o nome exato da tarefa E do projeto.")

        auth_headers = await get_api_auth_headers(client, use_json=True)
        
        allowed = {"nome", "descricao", "prioridade", "status", "prazo", "projeto_id"}
        payload = {k: v for k, v in patch.items() if k in allowed and v is not None}
        
        if "prazo" in patch and patch["prazo"]:
            payload["prazo"] = _parse_date_robust(patch["prazo"])

        if "responsavel" in patch and patch["responsavel"]:
            resp_id = await resolve_responsavel_id(client, patch["responsavel"], default_user_id=user_id)
            payload["responsavel_id"] = resp_id
        elif "responsavel_id" in patch and patch["responsavel_id"]:
             payload["responsavel_id"] = patch["responsavel_id"]
        
        if not payload: 
            raise ValueError("patch vazio")
            
        url = f"{TASKS_API_BASE}/tarefas/{resolved_tid}"
        resp = await client.put(url, json=payload, headers=auth_headers)
        resp.raise_for_status(); return sanitize_doc(resp.json())
        
def list_tasks_by_status(status: str, top_k: int = 50) -> List[Dict[str, Any]]:
    status_norm = (status or "").strip()
    if not status_norm: return []
    rx = {"$regex": f"^{re.escape(status_norm)}$", "$options": "i"}
    employee_map = _get_employee_map()
    tasks_raw = mongo()[COLL_TAREFAS].find({"status": rx}).sort("prazo", 1).limit(top_k)
    tasks_clean = [sanitize_doc(t) for t in tasks_raw]
    return [_enrich_doc_with_responsavel(t, employee_map) for t in tasks_clean]

def count_tasks_by_status(status: str) -> int:
    status_norm = (status or "").strip()
    if not status_norm: return 0
    rx = {"$regex": f"^{re.escape(status_norm)}$", "$options": "i"}
    try:
        return mongo()[COLL_TAREFAS].count_documents({"status": rx})
    except Exception as e:
        print(f"Erro ao contar tarefas por status: {e}")
        return -1

def find_project_responsavel(project_name: str) -> str:
    project_name_norm = (project_name or "").strip()
    if not project_name_norm:
        return "Nome do projeto não fornecido."
    rx = {"$regex": f"^{re.escape(project_name_norm)}$", "$options": "i"}
    proj = mongo()[COLL_PROJETOS].find_one({"nome": rx})
    if not proj:
        return f"Projeto '{project_name_norm}' não encontrado."
    proj_clean = sanitize_doc(proj)
    employee_map = _get_employee_map()
    enriched_proj = _enrich_doc_with_responsavel(proj_clean, employee_map) 
    return enriched_proj.get("responsavel_nome", "(Responsável não definido)")

def count_tasks_in_project(project_name: str) -> int:
    project_name_norm = (project_name or "").strip()
    if not project_name_norm: return -1
    try:
        rx_proj = {"$regex": f"^{re.escape(project_name_norm)}$", "$options": "i"}
        proj = mongo()[COLL_PROJETOS].find_one({"nome": rx_proj}, {"_id": 1})
        if not proj: return -2
        project_id = proj.get("_id")
        task_query = {"projeto": DBRef(collection=COLL_PROJETOS, id=project_id)}
        return mongo()[COLL_TAREFAS].count_documents(task_query)
    except Exception as e:
        print(f"Erro ao contar tarefas no projeto: {e}")
        return -3

def count_projects_by_responsavel(responsavel_id_str: Optional[str]) -> int:
    if not responsavel_id_str: return -1
    try:
        resp_oid = to_oid(responsavel_id_str)
        query = {"responsavel": DBRef(collection=COLL_FUNCIONARIOS, id=resp_oid)}
        return mongo()[COLL_PROJETOS].count_documents(query)
    except Exception as e:
        print(f"Erro ao contar projetos por responsável: {e}")
        return -2

def list_projects_by_responsavel(responsavel_id_str: Optional[str], top_k: int = 50) -> List[Dict[str, Any]]:
    if not responsavel_id_str: return []
    try:
        resp_oid = to_oid(responsavel_id_str)
        query = {"responsavel": DBRef(collection=COLL_FUNCIONARIOS, id=resp_oid)}
        employee_map = _get_employee_map()
        projects_raw = mongo()[COLL_PROJETOS].find(query).sort("prazo", 1).limit(top_k)
        projects_clean = [sanitize_doc(p) for p in projects_raw]
        return [_enrich_doc_with_responsavel(p, employee_map) for p in projects_clean]
    except Exception as e:
        print(f"Erro ao listar projetos por responsável: {e}")
        return []

def list_tasks_by_project_name(project_name: str, top_k: int = 10) -> List[Dict[str, Any]]:
    project_name_norm = (project_name or "").strip()
    if not project_name_norm: return []
    try:
        rx_proj = {"$regex": f"^{re.escape(project_name_norm)}$", "$options": "i"}
        proj = mongo()[COLL_PROJETOS].find_one({"nome": rx_proj}, {"_id": 1})
        if not proj: return []
        project_id = proj.get("_id")
        task_query = {"projeto": DBRef(collection=COLL_PROJETOS, id=project_id)}
        employee_map = _get_employee_map()
        tasks_raw = mongo()[COLL_TAREFAS].find(task_query).sort("prazo", 1).limit(top_k)
        tasks_clean = [sanitize_doc(t) for t in tasks_raw]
        return [_enrich_doc_with_responsavel(t, employee_map) for t in tasks_clean]
    except Exception as e:
        print(f"Erro ao listar tarefas do projeto: {e}")
        return []

async def import_project_from_url_tool(
    xlsx_url: str, 
    projeto_nome: str, 
    projeto_situacao: str, 
    projeto_prazo: str, 
    projeto_responsavel: str,
    projeto_descricao: Optional[str] = None,
    projeto_categoria: Optional[str] = None,
    user_id: Optional[str] = None
) -> Dict[str, Any]:
    effective_xlsx_url = xlsx_url
    if "docs.google.com/spreadsheets" in (xlsx_url or "").lower():
        sheet_id = extract_gsheet_id(xlsx_url)
        if sheet_id:
            effective_xlsx_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
    prazo_formatado = _parse_date_robust(projeto_prazo)
    return await tasks_from_xlsx_logic(
        projeto_id=None, projeto_nome=projeto_nome,
        user_id=user_id,
        create_project_flag=1, projeto_situacao=projeto_situacao,
        projeto_prazo=prazo_formatado, projeto_responsavel=projeto_responsavel,
        projeto_descricao=projeto_descricao, projeto_categoria=projeto_categoria,
        xlsx_url=effective_xlsx_url,
        file_bytes=None
    )

async def get_pdf_content_from_url_impl(url: str) -> str:
    try:
        pdf_bytes = fetch_pdf_bytes(url)
        return extract_full_pdf_text(pdf_bytes)
    except Exception as e:
        return f"Erro ao processar PDF da URL: {str(e)}"

async def solve_pdf_enigma_from_url_impl(url: str) -> str:
    try:
        pdf_bytes = fetch_pdf_bytes(url)
        full_text = extract_full_pdf_text(pdf_bytes)
        if not full_text:
            return "Não foi possível extrair texto do PDF."
        message = extract_hidden_message(full_text)
        return message if message else "Nenhuma mensagem secreta encontrada."
    except Exception as e:
        return f"Erro ao processar enigma do PDF: {str(e)}"

async def _get_first_pdf_url_from_xlsx_url(url: str) -> str:
    """
    Helper (V17): Baixa um XLSX/GoogleSheet de uma URL, lê,
    e retorna o *primeiro* link de PDF encontrado.
    """
    sheet_id = extract_gsheet_id(url)
    effective_xlsx_url = url
    if sheet_id:
        effective_xlsx_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
    
    try:
        xlsx_bytes = fetch_bytes(effective_xlsx_url) # Re-usa função existente
        df = xlsx_bytes_to_dataframe_preserving_hyperlinks(xlsx_bytes) # Re-usa função existente
    except Exception as e:
        raise ValueError(f"Não foi possível baixar ou ler o arquivo XLSX da URL: {str(e)}")

    if "Documento Referência" not in df.columns:
        raise ValueError("O XLSX não tem a coluna 'Documento Referência'.")
        
    for pdf_url in df["Documento Referência"]:
        if pdf_url and str(pdf_url).lower().startswith("http"):
            return str(pdf_url) # Retorna o primeiro link de PDF válido
            
    raise ValueError("Não encontrei nenhum link de PDF válido na coluna 'Documento Referência' do arquivo.")

async def get_pdf_content_from_xlsx_url_impl(url: str) -> str:
    """
    Tool Impl (V17): Pega URL de XLSX, acha URL de PDF dentro, e lê o texto.
    """
    try:
        pdf_url = await _get_first_pdf_url_from_xlsx_url(url)
        return await get_pdf_content_from_url_impl(pdf_url) # Re-usa função existente
    except Exception as e:
        return str(e)

async def solve_enigma_from_xlsx_url_impl(url: str) -> str:
    """
    Tool Impl (V17): Pega URL de XLSX, acha URL de PDF dentro, e resolve o enigma.
    """
    try:
        pdf_url = await _get_first_pdf_url_from_xlsx_url(url)
        return await solve_pdf_enigma_from_url_impl(pdf_url) # Re-usa função existente
    except Exception as e:
        return str(e)

def toolset() -> Tool:
    fns = [
        FunctionDeclaration(name="count_all_projects", description="Conta e retorna o número total de projetos na base de dados (IGNORA o contexto de funcionário). Use para 'quantos projetos existem?' ou 'qual o total de projetos?'.", parameters={"type": "object", "properties": {}}),        FunctionDeclaration(name="count_all_tasks", description="Conta e retorna o número total de tarefas.", parameters={"type": "object", "properties": {}}),
        FunctionDeclaration(name="count_projects_by_status", description="Conta e retorna o número de projetos por status (ex: 'em andamento').", parameters={"type": "object", "properties": {"status": {"type": "string"}}, "required": ["status"]}),
        FunctionDeclaration(name="list_all_projects", description="Lista todos os projetos.", parameters={"type": "object", "properties": {}}),
        FunctionDeclaration(name="list_all_tasks", description="Lista todas as tarefas.", parameters={"type": "object", "properties": {}}),
        FunctionDeclaration(name="list_all_funcionarios", description="Lista todos os funcionários.", parameters={"type": "object", "properties": {}}),
        FunctionDeclaration(name="list_tasks_by_deadline_range", description="Lista tarefas com prazo entre datas (YYYY-MM-DD).", parameters={"type": "object", "properties": {"start": {"type": "string"}, "end": {"type": "string"}}, "required": ["start", "end"]}),
        FunctionDeclaration(name="upcoming_deadlines", description="Lista tarefas com prazo vencendo nos próximos X dias.", parameters={"type": "object", "properties": {"days": {"type": "integer"}}, "required": ["days"]}),
        FunctionDeclaration(name="list_projects_by_status", description="Lista projetos por status (ex: 'em andamento').", parameters={"type": "object", "properties": {"status": {"type": "string"}}, "required": ["status"]}),
        FunctionDeclaration(name="count_my_projects", description="Conta quantos projetos são de responsabilidade do usuário ATUAL.", parameters={"type": "object", "properties": {}}),
        FunctionDeclaration(name="list_my_projects", description="Lista os projetos de responsabilidade do usuário ATUAL.", parameters={"type": "object", "properties": {}}),
        FunctionDeclaration(name="update_project", description="Atualiza campos de um projeto. Identifique o projeto usando 'project_id' (preferencial) OU 'project_name'.", parameters={"type": "object", "properties": {"patch": {"type": "object", "description": "Os campos para atualizar. Ex: {'situacao': 'concluído', 'prazo': '2025-12-31'}", "properties": {"nome": {"type": "string"}, "situacao": {"type": "string"}, "prazo": {"type": "string"}, "responsavel": {"type": "string", "description": "O NOME do novo responsável (ex: 'Ana Luiza Dourado')."}, "descricao": {"type": "string"}, "categoria": {"type": "string"}}}, "project_id": {"type": "string", "description": "O ID do projeto (se você já souber)."}, "project_name": {"type": "string", "description": "O NOME EXATO do projeto (usar se 'project_id' for desconhecido)."}}, "required": ["patch"]}),        
        FunctionDeclaration(name="create_project", description="Cria um novo projeto.", parameters={"type": "object", "properties": {"nome": {"type": "string"}, "responsavel": {"type": "string"}, "situacao": {"type": "string"}, "prazo": {"type": "string"}}, "required": ["nome", "responsavel", "situacao", "prazo"]}),
        FunctionDeclaration(name="create_task", description="Cria um nova tarefa.", parameters={"type": "object", "properties": {"nome": {"type": "string"}, "projeto_id": {"type": "string"}, "responsavel_id": {"type": "string"}, "prazo": {"type": "string"}, "status": {"type": "string"}}, "required": ["nome", "projeto_id", "responsavel_id", "prazo", "status"]}),
        FunctionDeclaration(name="update_task", description="Atualiza campos de uma tarefa. Você deve identificar a tarefa usando 'task_id' (preferencial) OU usando 'task_name' e 'project_name' juntos.", parameters={ "type": "object", "properties": { "patch": { "type": "object", "description": "Os campos para atualizar. Ex: {'status': 'concluída', 'prazo': '2025-12-31'}", "properties": { "nome": {"type": "string"}, "status": {"type": "string"}, "prazo": {"type": "string"}, "responsavel_id": {"type": "string", "description": "O ID do responsável (se souber)."}, "responsavel": {"type": "string", "description": "O NOME do responsável (ex: 'Ana Luiza Dourado')."}, "descricao": {"type": "string"}, "prioridade": {"type": "string"}, }}, "task_id": {"type": "string", "description": "O ID da tarefa (se você já souber)."}, "task_name": {"type": "string", "description": "O NOME EXATO da tarefa (usar se 'task_id' for desconhecido)."}, "project_name": {"type": "string", "description": "O NOME EXATO do projeto (OBRIGATÓRIO se 'task_name' for usado)."}}, "required": ["patch"]}),        
        FunctionDeclaration(name="import_project_from_url", description="Cria um projeto e importa tarefas a partir de uma URL de arquivo .xlsx ou Google Sheets.", parameters={"type": "object", "properties": {"xlsx_url": {"type": "string"}, "projeto_nome": {"type": "string"}, "projeto_situacao": {"type": "string"}, "projeto_prazo": {"type": "string"}, "projeto_responsavel": {"type": "string"}, "projeto_descricao": {"type": "string"}, "projeto_categoria": {"type": "string"}}, "required": ["xlsx_url", "projeto_nome", "projeto_situacao", "projeto_prazo", "projeto_responsavel"]}),
        FunctionDeclaration(name="list_tasks_by_status", description="Lista tarefas com base em um status exato (ex: 'não iniciada', 'concluída').", parameters={"type": "object", "properties": {"status": {"type": "string"}}, "required": ["status"]}),
        FunctionDeclaration(name="count_tasks_by_status", description="Conta tarefas com base em um status exato (ex: 'não iniciada', 'concluída').", parameters={"type": "object", "properties": {"status": {"type": "string"}}, "required": ["status"]}),
        FunctionDeclaration(name="find_project_responsavel", description="Encontra o nome do responsável por um projeto (busca por nome exato).", parameters={"type": "object", "properties": {"project_name": {"type": "string"}}, "required": ["project_name"]}),
        FunctionDeclaration(name="count_tasks_in_project", description="Conta o número de tarefas em um projeto (busca por nome exato).", parameters={"type": "object", "properties": {"project_name": {"type": "string"}}, "required": ["project_name"]}),
        FunctionDeclaration(name="list_tasks_by_project_name", description="Lista as N primeiras tarefas de um projeto (busca por nome exato).", parameters={"type": "object", "properties": {"project_name": {"type": "string"}, "top_k": {"type": "integer"}}, "required": ["project_name"]}),
        
        FunctionDeclaration(name="get_pdf_content_from_url", description="Extrai e retorna todo o texto de um arquivo PDF hospedado em uma URL. Use isso para 'ler' ou 'analisar' um PDF.", parameters={"type": "object", "properties": {"url": {"type": "string"}}, "required": ["url"]}),
        FunctionDeclaration(name="solve_pdf_enigma_from_url", description="Encontra uma 'frase secreta' escondida em um PDF (letras maiúsculas fora de lugar) a partir de uma URL.", parameters={"type": "object", "properties": {"url": {"type": "string"}}, "required": ["url"]}),

        FunctionDeclaration(name="find_employee_with_most_tasks", description="Encontra o funcionário (colaborador) que possui o maior número de tarefas atribuídas e retorna o nome do funcionário e a contagem total de tarefas.", parameters={"type": "object", "properties": {}}),
        FunctionDeclaration(name="find_employee_with_least_tasks", description="Encontra o funcionário (colaborador) com o MENOR número de tarefas atribuídas (mínimo 1).", parameters={"type": "object", "properties": {}}),
        FunctionDeclaration(name="find_employee_with_most_projects", description="Encontra o funcionário (colaborador) com o MAIOR número de projetos atribuídos.", parameters={"type": "object", "properties": {}}),
        FunctionDeclaration(name="find_employee_with_least_projects", description="Encontra o funcionário (colaborador) com o MENOR número de projetos atribuídos (mínimo 1).", parameters={"type": "object", "properties": {}}),
        FunctionDeclaration(name="count_tasks_by_employee_name", description="Conta o número total de tarefas de um funcionário específico (buscando por nome ou email).", parameters={"type": "object", "properties": {"employee_name": {"type": "string"}}, "required": ["employee_name"]}),
        FunctionDeclaration(name="count_projects_by_employee_name", description="Conta o número de projetos de um funcionário específico. Use APENAS se um nome de funcionário for mencionado na pergunta ou no contexto imediato.", parameters={"type": "object", "properties": {"employee_name": {"type": "string"}}, "required": ["employee_name"]}),
        
        FunctionDeclaration(name="get_employee_email", description="Encontra e retorna o email de um funcionário específico (buscando por nome).", parameters={"type": "object", "properties": {"employee_name": {"type": "string"}}, "required": ["employee_name"]}),
        FunctionDeclaration(name="get_employee_position", description="Encontra e retorna o cargo (posição) de um funcionário específico (buscando por nome).", parameters={"type": "object", "properties": {"employee_name": {"type": "string"}}, "required": ["employee_name"]}),
        FunctionDeclaration(name="get_employee_department", description="Encontra e retorna o departamento (área) de um funcionário específico (buscando por nome).", parameters={"type": "object", "properties": {"employee_name": {"type": "string"}}, "required": ["employee_name"]}),
        FunctionDeclaration(name="get_employee_tenure", description="Busca o tempo de casa E a data de admissão de um funcionário. A ferramenta retorna DOIS campos: 'tempo_de_casa' (ex: '1 mês e 21 dias') e 'data_admissao' (ex: '16/09/2025'). Use o campo apropriado para a pergunta do usuário.", parameters={"type": "object", "properties": {"employee_name": {"type": "string"}}, "required": ["employee_name"]}),
        
        FunctionDeclaration(name="get_pdf_content_from_xlsx_url", description="Extrai o texto de um PDF que está LINKADO DENTRO de um arquivo .xlsx ou Google Sheets. Use se o usuário fornecer uma URL de planilham, mas pedir para 'ler o PDF' ou 'resumir o PDF'.", parameters={"type": "object", "properties": {"url": {"type": "string"}}, "required": ["url"]}),
        FunctionDeclaration(name="solve_enigma_from_xlsx_url", description="Resolve um enigma de um PDF que está LINKADO DENTRO de um arquivo .xlsx ou Google Sheets. Use se o usuário fornecer uma URL de planilha e pedir o 'enigma' ou 'frase secreta'.", parameters={"type": "object", "properties": {"url": {"type": "string"}}, "required": ["url"]}),
    ]
    return Tool(function_declarations=fns)

async def exec_tool(name: str, args: Dict[str, Any], user_id: Optional[str] = None) -> Dict[str, Any]:
    try:
        if name == "count_all_projects": return {"ok": True, "data": count_all_projects()}
        if name == "count_all_tasks": return {"ok": True, "data": count_all_tasks()}
        if name == "count_projects_by_status": return {"ok": True, "data": count_projects_by_status(args["status"])}
        if name == "list_all_projects": return {"ok": True, "data": list_all_projects(args.get("top_k", 500))}
        if name == "list_all_tasks": return {"ok": True, "data": list_all_tasks(args.get("top_k", 2000))}
        if name == "list_all_funcionarios": return {"ok": True, "data": list_all_funcionarios(args.get("top_k", 500))}
        if name == "list_tasks_by_deadline_range": return {"ok": True, "data": list_tasks_by_deadline_range(args["start"], args["end"], args.get("top_k", 50))}
        if name == "upcoming_deadlines": return {"ok": True, "data": upcoming_deadlines(args.get("days", 14), args.get("top_k", 50))}
        if name == "count_my_projects": return {"ok": True, "data": count_projects_by_responsavel(user_id)}
        if name == "list_my_projects": return {"ok": True, "data": list_projects_by_responsavel(user_id, args.get("top_k", 50))}
        if name == "list_projects_by_status": return {"ok": True, "data": list_projects_by_status(args["status"], args.get("top_k", 50))}
        if name == "list_tasks_by_status": return {"ok": True, "data": list_tasks_by_status(args["status"], args.get("top_k", 50))}
        if name == "count_tasks_by_status": return {"ok": True, "data": count_tasks_by_status(args["status"])}
        if name == "find_project_responsavel": return {"ok": True, "data": find_project_responsavel(args["project_name"])}
        if name == "count_tasks_in_project": return {"ok": True, "data": count_tasks_in_project(args["project_name"])}
        if name == "list_tasks_by_project_name": return {"ok": True, "data": list_tasks_by_project_name(args["project_name"], args.get("top_k", 10))}
        if name == "update_project": return {"ok": True, "data": await update_project(patch=args.get("patch", {}), project_id=args.get("project_id"), project_name=args.get("project_name"), user_id=user_id)}              
        if name == "create_project": return {"ok": True, "data": await create_project(args, user_id=user_id)}
        if name == "create_task": return {"ok": True, "data": await create_task(args)}
        if name == "update_task": return {"ok": True, "data": await update_task(**args)}               
        if name == "import_project_from_url": return {"ok": True, "data": await import_project_from_url_tool(**args, user_id=user_id)}
        
        if name == "find_employee_with_most_tasks": return {"ok": True, "data": find_employee_with_most_tasks()}
        if name == "find_employee_with_least_tasks": return {"ok": True, "data": find_employee_with_least_tasks()}
        if name == "find_employee_with_most_projects": return {"ok": True, "data": find_employee_with_most_projects()}
        if name == "find_employee_with_least_projects": return {"ok": True, "data": find_employee_with_least_projects()}
        if name == "count_tasks_by_employee_name": return {"ok": True, "data": count_tasks_by_employee_name(args["employee_name"])}
        if name == "count_projects_by_employee_name": return {"ok": True, "data": count_projects_by_employee_name(args["employee_name"])}

        if name == "get_employee_email": return {"ok": True, "data": get_employee_email(args["employee_name"])}
        if name == "get_employee_position": return {"ok": True, "data": get_employee_position(args["employee_name"])}
        if name == "get_employee_department": return {"ok": True, "data": get_employee_department(args["employee_name"])}
        if name == "get_employee_tenure": return {"ok": True, "data": get_employee_tenure(args["employee_name"])}

        if name == "get_pdf_content_from_url": return {"ok": True, "data": await get_pdf_content_from_url_impl(args["url"])}
        if name == "solve_pdf_enigma_from_url": return {"ok": True, "data": await solve_pdf_enigma_from_url_impl(args["url"])}

        if name == "get_pdf_content_from_xlsx_url": return {"ok": True, "data": await get_pdf_content_from_xlsx_url_impl(args["url"])}
        if name == "solve_enigma_from_xlsx_url": return {"ok": True, "data": await solve_enigma_from_xlsx_url_impl(args["url"])}

        return {"ok": False, "error": f"função desconhecida: {name}"}
    except Exception as e:
        detail = str(e)
        if isinstance(e, httpx.HTTPStatusError):
            try: 
                err_json = e.response.json()
                detail = err_json.get("detail", err_json.get("erro", str(e)))
            except Exception: 
                detail = e.response.text
        return {"ok": False, "error": detail}

def _normalize_answer(raw: str, nome_usuario: str) -> str:
    raw = re.sub(r"[*_`#>]+", "", raw).strip()
    if all(sym not in raw for sym in ("🙂", "😊", "👋")):
        raw = raw.rstrip(".") + " 🙂"
    return raw
def init_model(system_instruction: str) -> GenerativeModel:
    vertex_init(project=PROJECT_ID, location=LOCATION) 
    return GenerativeModel(GEMINI_MODEL_ID, system_instruction=system_instruction)

async def chat_with_tools(user_msg: str, history: Optional[List[HistoryMessage]] = None, nome_usuario: Optional[str] = None, email_usuario: Optional[str] = None, id_usuario: Optional[str] = None) -> Dict[str, Any]:
    data_hoje, (inicio_mes, fim_mes) = iso_date(today()), month_bounds(today())
    nome_usuario = nome_usuario or "você"
    email_usuario = email_usuario or "email.desconhecido"
    id_usuario = id_usuario or "id.desconhecido"
    system_prompt_filled = SYSTEM_PROMPT.format(
        nome_usuario=nome_usuario, email_usuario=email_usuario, id_usuario=id_usuario,
        data_hoje=data_hoje, inicio_mes=inicio_mes, fim_mes=fim_mes,
    )
    model = init_model(system_prompt_filled)
    contents: List[Content] = []
    if history:
        for h in history:
            role_from_frontend = h.sender
            gemini_role = "model" if role_from_frontend == "ai" else "user"
            text_content = h.content.conteudo_texto
            contents.append(Content(role=gemini_role, parts=[Part.from_text(text_content)]))
    contents.append(Content(role="user", parts=[Part.from_text(user_msg)]))
    tools = [toolset()]
    tool_steps: List[Dict[str, Any]] = []
    for step in range(MAX_TOOL_STEPS):
        resp = model.generate_content(contents, tools=tools)
        calls = []
        model_response_content = None
        if resp.candidates and resp.candidates[0].content:
            model_response_content = resp.candidates[0].content
            if model_response_content.parts:
                for part in model_response_content.parts:
                    if getattr(part, "function_call", None): 
                        calls.append(part.function_call)

        if not calls:
            final_text = ""
            if resp.candidates and resp.candidates[0].content and resp.candidates[0].content.parts:
                final_text = getattr(resp.candidates[0].content.parts[0], "text", "") or ""
            final_text = re.sub(r"(?i)(aguarde( um instante)?|só um momento|apenas um instante)[^\n]*", "", final_text).strip()
            return {"answer": _normalize_answer(final_text, nome_usuario), "tool_steps": tool_steps}

        if model_response_content:
            contents.append(model_response_content) 
            
        for fc in calls:
            name = fc.name
            args = sanitize_doc(dict(fc.args or {}))
            if name in ("list_projects_by_deadline_range", "list_tasks_by_deadline_range") and (not args.get("start") or not args.get("end")):
                args["start"], args["end"] = inicio_mes, fim_mes
            result = await exec_tool(name, args, id_usuario)
            safe_result_for_log = sanitize_doc(result)
            safe_result_for_model = safe_result_for_log
            tool_steps.append({"call": {"name": name, "args": args}, "result": safe_result_for_log})
            contents.append(Content(role="tool", parts=[Part.from_function_response(name=name, response=safe_result_for_model)]))
            
    return {"answer": _normalize_answer("Concluí as ações solicitadas.", nome_usuario), "tool_steps": tool_steps}

async def handle_file_chat_from_context(req: ChatRequest, context_doc: Dict[str, Any]) -> JSONResponse:
    """
    Processa uma pergunta de chat contra um contexto de texto já salvo (do Mongo).
    Esta função contém a lógica de RAG e Enigma (Enigma APENAS para PDF).
    """
    pergunta = req.pergunta
    raw_file_text = context_doc.get("text_content", "")
    context_filename = context_doc.get("filename", "o arquivo")
    context_type = context_doc.get("context_type", "pdf") 
    
    nome_usuario_fmt = req.nome_usuario or "você"
    email_usuario_fmt = req.email_usuario or "email.desconcido"
    id_usuario_fmt = req.id_usuario or "id.desconhecido"

    final_answer = ""
    tool_steps = []

    try:
        pdf_enigma_tool_list = [
            FunctionDeclaration(
                name="solve_pdf_enigma",
                description="""CHAMADA OBRIGATÓRIA. Use esta ferramenta se a pergunta do usuário contiver QUALQUER palavra relacionada a: 'enigma', 'frase secreta', 'código secreto', 'mensagem escondida', 'frase escondida', 'código', 'decifrar', 'encontrar a frase'. Não use RAG se essas palavras estiverem presentes.""",
                parameters={"type": "object", "properties": {}}
            )
        ]
        pdf_tool = Tool(function_declarations=pdf_enigma_tool_list)

        xlsx_import_tool_list = [
            FunctionDeclaration(
                name="import_project_from_context",
                description="""CHAMADA OBRIGATÓRIA. Use esta ferramenta se a pergunta do usuário contiver intenção de 'importar', 'criar projeto', 'subir tarefas', 'adicionar da planilha' e o contexto for um arquivo XLSX/CSV. Você DEVE extrair todos os parâmetros solicitados.""",
                parameters={
                    "type": "object",
                    "properties": {
                        "projeto_nome": {"type": "string", "description": "O nome do novo projeto."},
                        "projeto_situacao": {"type": "string", "description": "A situação do projeto (ex: Em planejamento, Em andamento)."},
                        "projeto_prazo": {"type": "string", "description": "O prazo final do projeto, no formato DD-MM-AAAA."},
                        "projeto_responsavel": {"type": "string", "description": "O nome ou email do responsável pelo projeto."},
                        "projeto_descricao": {"type": "string", "description": "Uma breve descrição do projeto (opcional)."},
                        "projeto_categoria": {"type": "string", "description": "A categoria do projeto (opcional)."},
                    },
                    "required": ["projeto_nome", "projeto_situacao", "projeto_prazo", "projeto_responsavel"]
                }
            )
        ]
        xlsx_tool = Tool(function_declarations=xlsx_import_tool_list)

        tools_to_use = []
        file_format_desc = "PDF"
        rag_text = ""
        
        if context_type == "pdf" or context_type == "pdf_from_xlsx":
            tools_to_use.append(pdf_tool)
            rag_text = clean_pdf_text(raw_file_text)
            file_format_desc = "PDF"
        elif context_type == "xlsx":
            tools_to_use.append(xlsx_tool)
            rag_text = raw_file_text
            file_format_desc = "CSV (de um XLSX)"
        else:
            rag_text = raw_file_text

        contexto_prompt = f"""
        Você é um assistente. Use o CONTEÚDO DO DOCUMENTO abaixo (referente ao arquivo '{context_filename}', formato {file_format_desc}) para responder a PERGUNTA DO USUÁRIO.
        Responda APENAS com base no CONTEÚDO DO DOCUMENTO.

        ==================== CONTEÚDO DO DOCUMENTO ====================
        {rag_text[:10000]}
        ===============================================================

        PERGUNTA DO USUÁRIO: {pergunta}
        """

        data_hoje, (inicio_mes, fim_mes) = iso_date(today()), month_bounds(today())
        system_prompt_filled = SYSTEM_PROMPT.format(
            nome_usuario=nome_usuario_fmt, email_usuario=email_usuario_fmt, id_usuario=id_usuario_fmt,
            data_hoje=data_hoje, inicio_mes=inicio_mes, fim_mes=fim_mes,
        )
        model = init_model(system_prompt_filled)
        contents = [Content(role="user", parts=[Part.from_text(contexto_prompt)])]

        resp = model.generate_content(contents, tools=tools_to_use)

        function_call = None

        if (resp.candidates and
            resp.candidates[0].content and
            resp.candidates[0].content.parts):
            function_call = getattr(resp.candidates[0].content.parts[0], "function_call", None)
        
        if function_call and function_call.name == "solve_pdf_enigma":
            print(f"[Context-Chat] Usuário {id_usuario_fmt} pediu enigma do contexto.")
            tool_steps.append({"call": {"name": "solve_pdf_enigma", "args": {}}, "result": None})
            try:
                raw_result = extract_hidden_message(raw_file_text)
                if not raw_result:
                    result = "Nenhuma mensagem secreta encontrada."
                    final_answer = "Analisei o arquivo, mas não encontrei nenhuma frase secreta."
                else:
                    cleanup_prompt = f"""
                    Minha função Python extraiu a seguinte frase secreta literal: "{raw_result}"
                    
                    Sua tarefa é formatar esta frase para que ela se torne legível em português, seguindo regras MUITO ESTRITAS:
                    
                    1.  **Ordem e Letras:** A ordem das letras e as próprias letras NÃO PODEM MUDAR.
                        (Exemplo: O bloco 'EQ UI PE' deve se tornar 'EQUIPE'. O 'Q' não pode ser inventado. O 'E' inicial não pode virar 'É'.)
                    2.  **Acentuação:** Adicione a acentuação correta onde necessário. (Ex: 'VOCES' -> 'VOCÊS', 'SAO' -> 'SÃO').
                    3.  **Sem Pontuação:** NÃO adicione NENHUMA pontuação (vírgulas, pontos de exclamação, etc.). A saída deve ser "texto liso".
                    4.  **Caixa Alta:** A saída final deve ser TODA EM MAIÚSCULAS.
                    5.  **Palavras Estrangeiras:** PODE HAVER palavras estrangeiras (ex: 'TEAM', 'PROJECT', INNOVATION, etc.), mas mantenha a acentuação e caixa alta conforme as regras acima; se atente ao contexto para descobrir se é uma palavra estrangeira ou brasileira.
                    
                    Frase Bruta: "{raw_result}"
                    Transforme isso em uma frase limpa, em português, toda em maiúsculas, sem pontuação.
                    
                    Exemplo de 'antes' e 'depois':
                    - Antes: "EQ UI PE VO C E S PA SS AR AM"
                    - Depois: "EQUIPE VOCÊS PASSARAM"
                    
                    Responda APENAS com a frase final limpa.
                    """
                    
                    print("[DEBUG-V16-Context] Chamando IA (Gemini) para limpeza ESTRITA...")
                    cleanup_contents = [Content(role="user", parts=[Part.from_text(cleanup_prompt)])]
                    cleanup_resp = model.generate_content(cleanup_contents, tools=[])
                    cleaned_result = ""

                    if cleanup_resp.candidates and cleanup_resp.candidates[0].content and cleanup_resp.candidates[0].content.parts:
                        cleaned_result = getattr(cleanup_resp.candidates[0].content.parts[0], "text", "").strip()

                    if not cleaned_result or cleaned_result == raw_result:
                        result = raw_result
                        final_answer = f"A frase secreta encontrada no arquivo é: {raw_result}. Não consegui formatá-la melhor."
                    else:
                        result = cleaned_result
                        final_answer = f"A frase secreta encontrada no arquivo é: {cleaned_result}"
                tool_steps[-1]["result"] = result

            except Exception as e:
                detail = f"Erro ao processar enigma do contexto: {str(e)}"
                print(f"[Context-Chat] {detail}")
                final_answer = detail
                tool_steps[-1]["result"] = {"error": detail}

        elif function_call and function_call.name == "import_project_from_context":
            print(f"[Context-V19] Intenção: IMPORTAR (do contexto). UserID: {id_usuario_fmt}")
            tool_steps.append({"call": "import_project_from_context", "args": sanitize_doc(dict(function_call.args or {})), "result": None})
            
            try:
                params = function_call.args
                # 1. Obter os *bytes* do XLSX salvos no contexto
                xlsx_bytes_from_context = context_doc.get("binary_content")
                if not xlsx_bytes_from_context:
                    raise HTTPException(status_code=500, detail="Contexto XLSX encontrado, mas os dados binários (para importação) estão faltando. Por favor, anexe o arquivo novamente.")
                
                # 2. Validar prazo (copiado do /ai/chat-with-xlsx)
                prazo_fmt = _parse_date_robust(params.get("projeto_prazo"))

                # 3. Chamar a lógica de importação principal
                result = await tasks_from_xlsx_logic(
                    projeto_id=None,
                    projeto_nome=params.get("projeto_nome"),
                    user_id=id_usuario_fmt,
                    create_project_flag=1,
                    projeto_situacao=params.get("projeto_situacao"),
                    projeto_prazo=prazo_fmt,
                    projeto_responsavel=params.get("projeto_responsavel"),
                    projeto_descricao=params.get("projeto_descricao"),
                    projeto_categoria=params.get("projeto_categoria"),
                    xlsx_url=None,
                    file_bytes=xlsx_bytes_from_context
                )
                
                total = result.get('total', 0)
                erros = len(result.get('erros', []))
                final_answer = f"Projeto '{params.get('projeto_nome')}' criado com sucesso a partir do arquivo '{context_filename}'! Foram importadas {total} tarefas, com {erros} erros."
                tool_steps[-1]["result"] = result

            except Exception as e:
                detail = str(e)
                if isinstance(e, HTTPException): detail = e.detail
                elif isinstance(e, httpx.HTTPStatusError):
                    try: detail = e.response.json().get("detail", str(e))
                    except Exception: detail = e.response.text
                print(f"[Context-V19] FALHA ao importar do contexto: {detail}")
                final_answer = f"Tentei importar o projeto a partir do arquivo '{context_filename}', mas falhei: {detail}"
                tool_steps[-1]["result"] = {"error": detail}

        else:
            print(f"[Context-V19] Intenção: RAG (do contexto). UserID: {id_usuario_fmt}")
            if resp.candidates and resp.candidates[0].content and resp.candidates[0].content.parts:
                final_answer = getattr(resp.candidates[0].content.parts[0], "text", "Não encontrei a resposta no documento.")
            else:
                final_answer = "Não encontrei a resposta no documento."
            tool_steps.append({"call": "RAG_on_Context", "args": {"pergunta": pergunta}, "result": final_answer})

        final_answer_fmt = _normalize_answer(final_answer, nome_usuario_fmt)
        response_data = {
            "tipo_resposta": "TEXTO_CONTEXTO",
            "conteudo_texto": final_answer_fmt,
            "dados": tool_steps,
        }
        return JSONResponse(sanitize_doc(response_data))


    except Exception as e:
        detail = str(e)
        if isinstance(e, HTTPException): detail = e.detail
        print(f"[Context-V19] ERRO GERAL em handle_file_chat_from_context: {detail}")
        raise e

# =========================
# Rotas FastAPI
# =========================
@app.post("/ai/chat")
async def ai_chat(req: ChatRequest, _=Depends(require_api_key)):
    context_doc = None
    if req.id_usuario:
        try:
            db = mongo()
            context_doc = db[COLL_CONTEXTOS].find_one({"user_id": req.id_usuario})
        except Exception as e:
            print(f"[Context] Falha ao buscar contexto para {req.id_usuario}: {e}")

    if context_doc and context_doc.get("text_content"):
        print(f"[Context] Contexto encontrado para {req.id_usuario} (Tipo: {context_doc.get('context_type')}). Roteando para chat de PDF.")
        return await handle_file_chat_from_context(req, context_doc)
    print(f"[Context] Contexto não encontrado para {req.id_usuario}. Roteando para chat geral.")
    
    out = await chat_with_tools(
        user_msg=req.pergunta, 
        history=req.history, 
        nome_usuario=req.nome_usuario,
        email_usuario=req.email_usuario, 
        id_usuario=req.id_usuario        
    )
    response_data = {
        "tipo_resposta": "TEXTO",
        "conteudo_texto": out.get("answer", "Desculpe, não consegui processar sua solicitação."),
        "dados": out.get("tool_steps")
    }
    return JSONResponse(sanitize_doc(response_data))

@app.post("/tasks/from-xlsx")
async def tasks_from_xlsx(
    _=Depends(require_api_key), 
    projeto_id: Optional[str] = Form(None),
    projeto_nome: Optional[str] = Form(None),
    id_usuario: Optional[str] = Form(None),
    create_project_flag: int = Form(0),
    projeto_situacao: Optional[str] = Form(None),
    projeto_prazo: Optional[str] = Form(None),
    projeto_responsavel: Optional[str] = Form(None),
    projeto_descricao: Optional[str] = Form(None),
    projeto_categoria: Optional[str] = Form(None),
    xlsx_url: Optional[str] = Form(None),
    google_sheet_url: Optional[str] = Form(None),
    file: Optional[UploadFile] = File(None)
):
    file_bytes = await file.read() if file else None
    effective_xlsx_url = xlsx_url

    sources = sum(p is not None for p in [file_bytes, xlsx_url, google_sheet_url])
    
    if sources > 1:
        raise HTTPException(status_code=400, detail={"erro": "Use apenas UMA fonte de dados: 'google_sheet_url' OU 'xlsx_url' OU 'file'."})
    elif sources == 0:
        raise HTTPException(status_code=400, detail={"erro": "Nenhuma fonte de dados fornecida (google_sheet_url, xlsx_url, ou file)."})

    if google_sheet_url:
        sheet_id = extract_gsheet_id(google_sheet_url)
        if not sheet_id:
            raise HTTPException(status_code=400, detail={"erro": "Google Sheet URL inválida. Não foi possível extrair o ID."})
        
        effective_xlsx_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"

    result = await tasks_from_xlsx_logic(
        projeto_id=projeto_id, projeto_nome=projeto_nome,
        user_id=id_usuario,
        create_project_flag=create_project_flag, projeto_situacao=projeto_situacao,
        projeto_prazo=projeto_prazo, projeto_responsavel=projeto_responsavel,
        projeto_descricao=projeto_descricao, projeto_categoria=projeto_categoria,
        xlsx_url=effective_xlsx_url,
        file_bytes=file_bytes
    )
    return result

@app.post("/ai/chat-with-pdf")
async def ai_chat_with_pdf(
    pergunta: str = Form(...),
    file: UploadFile = File(...),
    nome_usuario: Optional[str] = Form(None),
    email_usuario: Optional[str] = Form(None),
    id_usuario: Optional[str] = Form(None),
    _ = Depends(require_api_key)
):
    """
    Endpoint de chat (V15 - Lógica Híbrida + Limpeza de IA)
    
    1. IA (Ferramenta) detecta a INTENÇÃO de "enigma".
    2. Python (extract_hidden_message) faz a EXTRAÇÃO LÓGICA (bruta).
    3. IA (2ª chamada) "limpa" o resultado do Python (acentos, espaços).
    
    MODIFICADO: Agora também salva o 'raw_pdf_text' no COLL_CONTEXTOS
    para persistência da sessão.
    """
    try:
        pdf_bytes = await file.read()
        raw_pdf_text = extract_full_pdf_text(pdf_bytes)
        
        data_hoje, (inicio_mes, fim_mes) = iso_date(today()), month_bounds(today())
        nome_usuario_fmt = nome_usuario or "você"
        email_usuario_fmt = email_usuario or "email.desconhecido"
        id_usuario_fmt = id_usuario
        if not id_usuario_fmt:
            id_usuario_fmt = "id.desconhecido.pdf" # Fallback
            print("[Context] ATENÇÃO: id_usuario não fornecido para /ai/chat-with-pdf. A memória não funcionará corretamente.")

        try:
            db = mongo()
            db[COLL_CONTEXTOS].update_one(
                {"user_id": id_usuario_fmt},
                {"$set": {
                    "user_id": id_usuario_fmt,
                    "context_type": "pdf",
                    "text_content": raw_pdf_text, # Salva o texto bruto (bom para enigma)
                    "filename": file.filename,
                    "updated_at": today()
                }},
                upsert=True
            )
            print(f"[Context] Contexto PDF salvo para {id_usuario_fmt} (Arquivo: {file.filename}).")
        except Exception as e:
            print(f"[Context] FALHA ao salvar contexto PDF para {id_usuario_fmt}: {e}")

        pdf_tools_list = [
            FunctionDeclaration(
                name="solve_pdf_enigma",
                description="""CHAMADA OBRIGATÓRIA. Use esta ferramenta se a pergunta do usuário contiver QUALQUER palavra relacionada a: 'enigma', 'frase secreta', 'código secreto', 'mensagem escondida', 'frase escondida', 'código', 'decifrar', 'encontrar a frase'. Não use RAG se essas palavras estiverem presentes.""",
                parameters={"type": "object", "properties": {}}
            )
        ]
        pdf_tool = Tool(function_declarations=pdf_tools_list)
        rag_text = clean_pdf_text(raw_pdf_text) 
        contexto_prompt = f"""
        Você é um assistente. Use o CONTEÚDO DO DOCUMENTO abaixo para responder a PERGUNTA DO USUÁRIO.
        Responda APENAS com base no CONTEÚDO DO DOCUMENTO.

        ==================== CONTEÚDO DO DOCUMENTO ====================
        {rag_text[:10000]} 
        ===============================================================

        PERGUNTA DO USUÁRIO: {pergunta}
        """
                
        system_prompt_filled = SYSTEM_PROMPT.format(
            nome_usuario=nome_usuario_fmt, email_usuario=email_usuario_fmt, id_usuario=id_usuario_fmt,
            data_hoje=data_hoje, inicio_mes=inicio_mes, fim_mes=fim_mes,
        )
        
        model = init_model(system_prompt_filled)
        contents = [Content(role="user", parts=[Part.from_text(contexto_prompt)])]
        
        resp = model.generate_content(contents, tools=[pdf_tool])
        
        if (
            resp.candidates and resp.candidates[0].content and 
            resp.candidates[0].content.parts and 
            getattr(resp.candidates[0].content.parts[0], "function_call", None)
        ):
            call = resp.candidates[0].content.parts[0].function_call
            
            if call.name == "solve_pdf_enigma":
                print("[DEBUG-V14] Ferramenta 'solve_pdf_enigma' foi chamada pela IA.")
                
                secret_message_raw = extract_hidden_message(raw_pdf_text)
                
                if not secret_message_raw:
                    print("[DEBUG-V15] Python (V14) não encontrou a frase.")
                    final_answer = "Analisei o arquivo, mas não encontrei nenhuma frase secreta."
                    final_answer = _normalize_answer(final_answer, nome_usuario_fmt)
                    
                    response_data = {
                        "tipo_resposta": "TEXTO_PDF",
                        "conteudo_texto": final_answer,
                        "dados": [
                            {
                                "call": {"name": "solve_pdf_enigma", "args": {}},
                                "result": {"status": "OK", "raw": secret_message_raw, "cleaned": final_answer_cleaned}
                            }
                        ]
                    }
                    return JSONResponse(sanitize_doc(response_data))


                print(f"[DEBUG-V15] Python (V14) encontrou: {secret_message_raw}")
                
                cleanup_prompt = f"""
                Minha função Python extraiu a seguinte frase secreta literal: "{secret_message_raw}"
                
                Sua tarefa é formatar esta frase para que ela se torne legível em português, seguindo regras MUITO ESTRITAS:
                
                1.  **Ordem e Letras:** A ordem das letras e as próprias letras NÃO PODEM MUDAR.
                    (Exemplo: O bloco 'EQ UI PE' deve se tornar 'EQUIPE'. O 'Q' não pode ser inventado. O 'E' inicial não pode virar 'É'.)
                2.  **Acentuação:** Adicione a acentuação correta onde necessário. (Ex: 'VOCES' -> 'VOCÊS', 'SAO' -> 'SÃO').
                3.  **Sem Pontuação:** NÃO adicione NENHUMA pontuação (vírgulas, pontos de exclamação, etc.). A saída deve ser "texto liso".
                4.  **Caixa Alta:** A saída final deve ser TODA EM MAIÚSCULAS.
                5.  **Palavras Estrangeiras:** PODE HAVER palavras estrangeiras (ex: 'TEAM', 'PROJECT', INNOVATION, etc.), mas mantenha a acentuação e caixa alta conforme as regras acima; se atente ao contexto para descobrir se é uma palavra estrangeira ou brasileira.
                
                Frase Bruta: "{secret_message_raw}"
                Transforme isso em uma frase limpa, em português, toda em maiúsculas, sem pontuação.
                
                Exemplo de 'antes' e 'depois':
                - Antes: "EQ UI PE VO C E S PA SS AR AM"
                - Depois: "EQUIPE VOCÊS PASSARAM"
                
                Responda APENAS com a frase final limpa.
                """
                
                print("[DEBUG-V16] Chamando IA (Gemini) para limpeza ESTRITA (Lógica Inline)...")
                
                cleanup_contents = [Content(role="user", parts=[Part.from_text(cleanup_prompt)])]
                cleanup_resp = model.generate_content(cleanup_contents, tools=[])
                                
                final_answer_cleaned = ""
                if cleanup_resp.candidates and cleanup_resp.candidates[0].content and cleanup_resp.candidates[0].content.parts:
                    final_answer_cleaned = getattr(cleanup_resp.candidates[0].content.parts[0], "text", "").strip()

                if not final_answer_cleaned or final_answer_cleaned == secret_message_raw:
                    print("[DEBUG-V15] IA de limpeza (Inline) falhou. Retornando frase bruta.")
                    final_answer = f"A frase secreta encontrada no arquivo é: {secret_message_raw}"
                else:
                    print(f"[DEBUG-V15] IA de limpeza (Inline) retornou: {final_answer_cleaned}")
                    final_answer = f"A frase secreta encontrada no arquivo é: {final_answer_cleaned}"

                final_answer = _normalize_answer(final_answer, nome_usuario_fmt)
                
                response_data = {
                    "tipo_resposta": "TEXTO_PDF",
                    "conteudo_texto": final_answer,
                    "dados": [
                        {
                            "call": {"name": "solve_pdf_enigma", "args": {}},
                            "result": {"status": "OK", "raw": secret_message_raw, "cleaned": final_answer_cleaned}
                        }
                    ]
                }
                return JSONResponse(sanitize_doc(response_data))
        
        print("[DEBUG-V14] Ferramenta de enigma não foi chamada. Usando RAG.")
        final_text = ""
        if resp.candidates and resp.candidates[0].content and resp.candidates[0].content.parts:
            final_text = getattr(resp.candidates[0].content.parts[0], "text", "") or ""

        if not rag_text:
             final_text = "Desculpe, não consegui ler o texto desse PDF para responder sua pergunta."
        elif not final_text:
             final_text = "Desculpe, não consegui processar sua solicitação sobre o PDF."

        final_answer = _normalize_answer(final_text, nome_usuario_fmt)
        
        response_data = {
            "tipo_resposta": "TEXTO_PDF",
            "conteudo_texto": final_answer,
            "dados": [
                {
                    "call": {"name": "RAG_simples", "args": {"pergunta": pergunta}},
                    "result": {"status": "OK", "answer": final_answer}
                }
            ]
        }
        return JSONResponse(sanitize_doc(response_data))

    except Exception as e:
        raise e

@app.post("/ai/chat-with-pdf-url")
async def ai_chat_with_pdf_url(
    req: ChatWithPdfUrlRequest,
    _ = Depends(require_api_key)
):
    """
    Endpoint de chat (V15 - Lógica Híbrida + Limpeza de IA) - VERSÃO URL
    
    Funciona exatamente como /ai/chat-with-pdf, mas baixa o PDF de uma URL
    usando a robusta função fetch_pdf_bytes().
    """
    try:
        print(f"[DEBUG-URL] Baixando PDF de: {req.pdf_url}")
        pdf_bytes = fetch_pdf_bytes(req.pdf_url)
        raw_pdf_text = extract_full_pdf_text(pdf_bytes)
        
        parsed_url = urlparse(req.pdf_url)
        filename_from_url = parsed_url.path.split('/')[-1] or "arquivo_da_url.pdf"
        
        data_hoje, (inicio_mes, fim_mes) = iso_date(today()), month_bounds(today())
        nome_usuario_fmt = req.nome_usuario or "você"
        email_usuario_fmt = req.email_usuario or "email.desconhecido"
        id_usuario_fmt = req.id_usuario
        
        if not id_usuario_fmt:
            id_usuario_fmt = "id.desconhecido.pdf-url"
            print("[Context] ATENÇÃO: id_usuario não fornecido para /ai/chat-with-pdf-url. A memória não funcionará corretamente.")

        try:
            db = mongo()
            db[COLL_CONTEXTOS].update_one(
                {"user_id": id_usuario_fmt},
                {"$set": {
                    "user_id": id_usuario_fmt,
                    "context_type": "pdf",
                    "text_content": raw_pdf_text,
                    "filename": filename_from_url,
                    "source_url": req.pdf_url,
                    "updated_at": today()
                }},
                upsert=True
            )
            print(f"[Context] Contexto PDF (de URL) salvo para {id_usuario_fmt} (Arquivo: {filename_from_url}).")
        except Exception as e:
            print(f"[Context] FALHA ao salvar contexto PDF (de URL) para {id_usuario_fmt}: {e}")
        
        pergunta_para_rag = re.sub(
            r"https?:\/\/[^\s]+", 
            "(o documento fornecido no contexto)", 
            req.pergunta, 
            flags=re.IGNORECASE
        ).strip()
        
        print(f"[DEBUG-URL-CLEAN] Pergunta Original: '{req.pergunta}'")
        print(f"[DEBUG-URL-CLEAN] Pergunta para RAG: '{pergunta_para_rag}'")

        pdf_tools_list = [
            FunctionDeclaration(
                name="solve_pdf_enigma",
                description="""CHAMADA OBRIGATÓRIA. Use esta ferramenta se a pergunta do usuário contiver QUALQUER palavra relacionada a: 'enigma', 'frase secreta', 'código secreto', 'mensagem escondida', 'frase escondida', 'código', 'decifrar', 'encontrar a frase'. Não use RAG se essas palavras estiverem presentes.""",
                parameters={"type": "object", "properties": {}}
            )
        ]
        pdf_tool = Tool(function_declarations=pdf_tools_list)
        rag_text = clean_pdf_text(raw_pdf_text) 
        contexto_prompt = f"""
        Você é um assistente. Use o CONTEÚDO DO DOCUMENTO abaixo para responder a PERGUNTA DO USUÁRIO.
        Responda APENAS com base no CONTEÚDO DO DOCUMENTO.

        ==================== CONTEÚDO DO DOCUMENTO ====================
        {rag_text[:10000]} 
        ===============================================================

        PERGUNTA DO USUÁRIO: {pergunta_para_rag}
        """
                
        system_prompt_filled = SYSTEM_PROMPT.format(
            nome_usuario=nome_usuario_fmt, email_usuario=email_usuario_fmt, id_usuario=id_usuario_fmt,
            data_hoje=data_hoje, inicio_mes=inicio_mes, fim_mes=fim_mes,
        )
        
        model = init_model(system_prompt_filled)
        contents = [Content(role="user", parts=[Part.from_text(contexto_prompt)])]
        
        resp = model.generate_content(contents, tools=[pdf_tool])
        
        if (
            resp.candidates and resp.candidates[0].content and 
            resp.candidates[0].content.parts and 
            getattr(resp.candidates[0].content.parts[0], "function_call", None)
        ):
            call = resp.candidates[0].content.parts[0].function_call
            
            if call.name == "solve_pdf_enigma":
                print("[DEBUG-V14-URL] Ferramenta 'solve_pdf_enigma' foi chamada pela IA (via URL).")
                
                secret_message_raw = extract_hidden_message(raw_pdf_text)
                
                if not secret_message_raw:
                    print("[DEBUG-V15-URL] Python (V14) não encontrou a frase.")
                    final_answer = "Analisei o arquivo, mas não encontrei nenhuma frase secreta."
                    final_answer = _normalize_answer(final_answer, nome_usuario_fmt)
                    
                    response_data = {
                        "tipo_resposta": "TEXTO_PDF",
                        "conteudo_texto": final_answer,
                        "dados": [
                            {
                                "call": {"name": "solve_pdf_enigma", "args": {}},
                                "result": {"status": "OK", "raw": secret_message_raw, "cleaned": None}
                            }
                        ]
                    }
                    return JSONResponse(sanitize_doc(response_data))


                print(f"[DEBUG-V15-URL] Python (V14) encontrou: {secret_message_raw}")
                
                cleanup_prompt = f"""
                Minha função Python extraiu a seguinte frase secreta literal: "{secret_message_raw}"
                
                Sua tarefa é formatar esta frase para que ela se torne legível em português, seguindo regras MUITO ESTRITAS:
                
                1.  **Ordem e Letras:** A ordem das letras e as próprias letras NÃO PODEM MUDAR.
                    (Exemplo: O bloco 'EQ UI PE' deve se tornar 'EQUIPE'. O 'Q' não pode ser inventado. O 'E' inicial não pode virar 'É'.)
                2.  **Acentuação:** Adicione a acentuação correta onde necessário. (Ex: 'VOCES' -> 'VOCÊS', 'SAO' -> 'SÃO').
                3.  **Sem Pontuação:** NÃO adicione NENHUMA pontuação (vírgulas, pontos de exclamação, etc.). A saída deve ser "texto liso".
                4.  **Caixa Alta:** A saída final deve ser TODA EM MAIÚSCULAS.
                5.  **Palavras Estrangeiras:** PODE HAVER palavras estrangeiras (ex: 'TEAM', 'PROJECT', INNOVATION, etc.), mas mantenha a acentuação e caixa alta conforme as regras acima; se atente ao contexto para descobrir se é uma palavra estrangeira ou brasileira.
                
                Frase Bruta: "{secret_message_raw}"
                Transforme isso em uma frase limpa, em português, toda em maiúsculas, sem pontuação.
                
                Exemplo de 'antes' e 'depois':
                - Antes: "EQ UI PE VO C E S PA SS AR AM"
                - Depois: "EQUIPE VOCÊS PASSARAM"
                
                Responda APENAS com a frase final limpa.
                """
                
                print("[DEBUG-V16-URL] Chamando IA (Gemini) para limpeza ESTRITA (Lógica Inline)...")
                
                cleanup_contents = [Content(role="user", parts=[Part.from_text(cleanup_prompt)])]
                cleanup_resp = model.generate_content(cleanup_contents, tools=[])
                                
                final_answer_cleaned = ""
                if cleanup_resp.candidates and cleanup_resp.candidates[0].content and cleanup_resp.candidates[0].content.parts:
                    final_answer_cleaned = getattr(cleanup_resp.candidates[0].content.parts[0], "text", "").strip()

                if not final_answer_cleaned or final_answer_cleaned == secret_message_raw:
                    print("[DEBUG-V15-URL] IA de limpeza (Inline) falhou. Retornando frase bruta.")
                    final_answer = f"A frase secreta encontrada no arquivo é: {secret_message_raw}"
                else:
                    print(f"[DEBUG-V15-URL] IA de limpeza (Inline) retornou: {final_answer_cleaned}")
                    final_answer = f"A frase secreta encontrada no arquivo é: {final_answer_cleaned}"

                final_answer = _normalize_answer(final_answer, nome_usuario_fmt)
                
                response_data = {
                    "tipo_resposta": "TEXTO_PDF",
                    "conteudo_texto": final_answer,
                    "dados": [
                        {
                            "call": {"name": "solve_pdf_enigma", "args": {}},
                            "result": {"status": "OK", "raw": secret_message_raw, "cleaned": final_answer_cleaned}
                        }
                    ]
                }
                return JSONResponse(sanitize_doc(response_data))
        
        print("[DEBUG-V14-URL] Ferramenta de enigma não foi chamada. Usando RAG (via URL).")
        final_text = ""
        if resp.candidates and resp.candidates[0].content and resp.candidates[0].content.parts:
            final_text = getattr(resp.candidates[0].content.parts[0], "text", "") or ""

        if not rag_text:
             final_text = "Desculpe, não consegui ler o texto desse PDF para responder sua pergunta."
        elif not final_text:
             final_text = "Desculpe, não consegui processar sua solicitação sobre o PDF."

        final_answer = _normalize_answer(final_text, nome_usuario_fmt)
        
        response_data = {
            "tipo_resposta": "TEXTO_PDF",
            "conteudo_texto": final_answer,
            "dados": [
                {
                    "call": {"name": "RAG_simples", "args": {"pergunta": req.pergunta}},
                    "result": {"status": "OK", "answer": final_answer}
                }
            ]
        }
        return JSONResponse(sanitize_doc(response_data))

    except Exception as e:
        raise e

@app.post("/ai/chat-with-xlsx")
async def ai_chat_with_xlsx(
    pergunta: str = Form(...),
    file: UploadFile = File(...),
    nome_usuario: Optional[str] = Form(None),
    email_usuario: Optional[str] = Form(None),
    id_usuario: Optional[str] = Form(None),
    _ = Depends(require_api_key)
):
    """
    Endpoint de chat (V17 - Lógica Híbrida com XLSX)
    
    1. Recebe um XLSX e uma pergunta.
    2. Identifica a intenção: "import", "enigma", ou "rag".
    3. Se "import", tenta extrair dados do projeto da pergunta e criar o projeto.
    4. Se "enigma" ou "rag", identifica a TAREFA, acha o link do PDF,
       E SALVA O TEXTO DO PDF NO CONTEXTO.
    """
    try:
        xlsx_bytes = await file.read()
        df = xlsx_bytes_to_dataframe_preserving_hyperlinks(xlsx_bytes)
        
        id_usuario_fmt = id_usuario
        if not id_usuario_fmt:
            id_usuario_fmt = "id.desconhecido.xlsx" # Fallback
            print("[Context] ATENÇÃO: id_usuario não fornecido para /ai/chat-with-xlsx. A memória não funcionará corretamente.")
        
        nome_usuario_fmt = nome_usuario or "você"
        email_usuario_fmt = email_usuario or "email.desconhecido"
        
        if "Nome" not in df.columns or "Documento Referência" not in df.columns:
            pass
        task_names = []
        if "Nome" in df.columns:
            task_names = df["Nome"].dropna().unique().tolist()
            task_map_str = "\n".join(f"- {name}" for name in task_names)
        else:
            task_map_str = "(Nenhuma coluna 'Nome' encontrada na planilha)"
            
        model = init_model("Você é um assistente de roteamento.")
        prompt_identificacao = f"""
        O usuário fez a pergunta: "{pergunta}"
        Ele também anexou um arquivo XLSX.
        O XLSX contém (potencialmente) estas tarefas:
        {task_map_str}
        Analise a pergunta e me retorne um JSON com:
        1. "task_name": O nome exato da tarefa da lista que o usuário mencionou (ou null se não for claro ou se a intenção for 'importar').
        2. "intention": "enigma" (se a pergunta for sobre enigma, frase secreta, etc.), "import" (se a pergunta for sobre 'importar', 'criar projeto', 'subir tarefas', 'adicionar da planilha'), ou "rag" (para qualquer outra pergunta, como resumir ou buscar).
        
        Responda APENAS com o JSON.
        """
        resp = model.generate_content([prompt_identificacao], tools=[])
        response_text = "{}".strip()
        if resp.candidates and resp.candidates[0].content and resp.candidates[0].content.parts:
            response_text = getattr(resp.candidates[0].content.parts[0], "text", "{}").replace("`", "").replace("json", "").strip()
        try:
            route_info = json.loads(response_text)
            task_name = route_info.get("task_name")
            intention = route_info.get("intention", "rag")
        except Exception:
            raise HTTPException(status_code=400, detail="Não consegui entender sua intenção (importar, ler, etc.).")
            
        final_answer = ""
        tool_steps = []
        
        if intention == "import":
            print(f"[DEBUG-V17] Intenção: IMPORT. Pergunta: {pergunta}")
            extract_prompt = f"""
            O usuário quer importar um projeto usando o arquivo XLSX anexado.
            Extraia os seguintes parâmetros da pergunta dele: "{pergunta}"
            
            - "projeto_nome": (string)
            - "projeto_situacao": (string)
            - "projeto_prazo": (string, formato DD-MM-AAAA)
            - "projeto_responsavel": (string, nome ou email)
            - "projeto_descricao": (string, opcional)
            - "projeto_categoria": (string, opcional)
            
            Responda APENAS com um JSON com os valores que encontrar. Use null se não encontrar.
            """
            extract_resp = model.generate_content([extract_prompt], tools=[])
            params_json_str = "{}"
            if extract_resp.candidates and extract_resp.candidates[0].content and extract_resp.candidates[0].content.parts:
                params_json_str = getattr(extract_resp.candidates[0].content.parts[0], "text", "{}").replace("`", "").replace("json", "").strip()
            try:
                params = json.loads(params_json_str)
            except Exception:
                params = {}
            required = {"projeto_nome", "projeto_situacao", "projeto_prazo", "projeto_responsavel"}
            missing = [r for r in required if not params.get(r)]          
            if missing:
                missing_str = ", ".join(missing)
                final_answer = f"Entendi que você quer importar esse arquivo! Para importar, preciso que você me envie o arquivo E, na *mesma* mensagem, me diga o nome do projeto, situação, prazo (DD-MM-AAAA) e o responsável. Campos faltando: {missing_str}."
                tool_steps.append({"call": "import_from_file", "args": params, "result": {"error": "missing_fields", "missing": missing}})
            else:
                try:
                    prazo_fmt = _parse_date_robust(params.get("projeto_prazo"))
                    result = await tasks_from_xlsx_logic(
                        projeto_id=None,
                        projeto_nome=params.get("projeto_nome"),
                        user_id=id_usuario, # Passa o id_usuario original
                        create_project_flag=1,
                        projeto_situacao=params.get("projeto_situacao"),
                        projeto_prazo=prazo_fmt,
                        projeto_responsavel=params.get("projeto_responsavel"),
                        projeto_descricao=params.get("projeto_descricao"),
                        projeto_categoria=params.get("projeto_categoria"),
                        xlsx_url=None,
                        file_bytes=xlsx_bytes
                    )                    
                    total = result.get('total', 0)
                    erros = len(result.get('erros', []))
                    final_answer = f"Projeto '{params.get('projeto_nome')}' criado com sucesso a partir do arquivo! Foram importadas {total} tarefas, com {erros} erros."
                    tool_steps.append({"call": "import_from_file", "args": params, "result": result})                
                except Exception as e:
                    detail = str(e)
                    if isinstance(e, HTTPException): detail = e.detail
                    elif isinstance(e, httpx.HTTPStatusError):
                        try: detail = e.response.json().get("detail", str(e))
                        except Exception: detail = e.response.text
                    final_answer = f"Tentei importar o projeto, mas falhei: {detail}"
                    tool_steps.append({"call": "import_from_file", "args": params, "result": {"error": detail}})    
        
        elif intention == "rag" and not task_name:
            print(f"[DEBUG-V19] Intenção: RAG GERAL (sobre o XLSX). Pergunta: {pergunta}")
            
            # 1. Converter DF para texto (CSV)
            xlsx_text_content = df.to_csv(index=False)
            if not xlsx_text_content.strip():
                 raise HTTPException(status_code=400, detail="O arquivo XLSX parece estar vazio ou não pude lê-lo.")

            # 2. Salvar no Contexto (para follow-ups)
            try:
                db = mongo()
                db[COLL_CONTEXTOS].update_one(
                    {"user_id": id_usuario_fmt},
                    {"$set": {
                        "user_id": id_usuario_fmt,
                        "context_type": "xlsx",
                        "text_content": xlsx_text_content, # (CSV) Para RAG futuro
                        "binary_content": xlsx_bytes,    # (Bytes) Para IMPORTAÇÃO futura
                        "filename": file.filename,
                        "updated_at": today()
                    }},
                    upsert=True
                )
                print(f"[Context-V19] Contexto XLSX GERAL (Texto+Binário) salvo para {id_usuario_fmt} (Arquivo: {file.filename}).")
            except Exception as e:
                print(f"[Context-V19] FALHA ao salvar contexto XLSX GERAL para {id_usuario_fmt}: {e}")

            # 3. Preparar RAG Prompt
            rag_prompt = f"""
            Use o CONTEÚDO DO DOCUMENTO XLSX abaixo (formatado como CSV) para responder a PERGUNTA DO USUÁRIO.
            Responda APENAS com base no CONTEÚDO DO DOCUMENTO.

            ==================== CONTEÚDO DO DOCUMENTO (CSV) ====================
            {xlsx_text_content[:10000]} 
            =====================================================================

            PERGUNTA DO USUÁRIO: {pergunta}
            """
            
            # 4. Chamar Gemini (reutilizando a lógica do RAG de PDF)
            data_hoje, (inicio_mes, fim_mes) = iso_date(today()), month_bounds(today())
            system_prompt_filled = SYSTEM_PROMPT.format(
                nome_usuario=nome_usuario_fmt, 
                email_usuario=email_usuario_fmt, 
                id_usuario=id_usuario_fmt,
                data_hoje=data_hoje, inicio_mes=inicio_mes, fim_mes=fim_mes,
            )
            rag_model = init_model(system_prompt_filled)
            rag_resp = rag_model.generate_content([rag_prompt], tools=[])
            
            if rag_resp.candidates and rag_resp.candidates[0].content and rag_resp.candidates[0].content.parts:
                final_answer = getattr(rag_resp.candidates[0].content.parts[0], "text", "Não encontrei a resposta no documento.")
            else:
                final_answer = "Não encontrei a resposta no documento."
            tool_steps.append({"call": "RAG_on_XLSX_General", "args": {"pergunta": pergunta}, "result": final_answer})
        
        elif not task_names:
            raise HTTPException(status_code=400, detail="A intenção não era 'importar', mas a planilha não tem uma coluna 'Nome' para que eu possa ler as tarefas.")
        elif not task_name or task_name not in task_names:
            raise HTTPException(status_code=404, detail=f"Não consegui identificar uma tarefa válida da sua planilha na sua pergunta. A intenção era 'ler' ou 'enigma'?")
        
        else:
            if "Documento Referência" not in df.columns:
                raise HTTPException(status_code=400, detail="O XLSX precisa da coluna 'Documento Referência' para RAG/Enigma.")
            task_row = df[df['Nome'] == task_name].iloc[0]
            pdf_url = task_row.get("Documento Referência")
            if not pdf_url or not str(pdf_url).lower().startswith("http"):
                raise HTTPException(status_code=404, detail=f"A tarefa '{task_name}' não possui um link de PDF válido no arquivo.")
            
            print(f"[DEBUG-V17] Intenção: {intention}. PDF: {pdf_url}")
            try:
                pdf_content_raw = await get_pdf_content_from_url_impl(pdf_url)
                if "Erro ao processar PDF" in pdf_content_raw or not pdf_content_raw.strip():
                    raise ValueError(pdf_content_raw)
            except Exception as e:
                raise HTTPException(status_code=500, detail=f"Erro ao ler o PDF da tarefa '{task_name}' (URL: {pdf_url}): {str(e)}")

            try:
                db = mongo()
                db[COLL_CONTEXTOS].update_one(
                    {"user_id": id_usuario_fmt},
                    {"$set": {
                        "user_id": id_usuario_fmt,
                        "context_type": "pdf_from_xlsx",
                        "text_content": pdf_content_raw, # Salva o texto bruto
                        "filename": file.filename, # Nome do XLSX original
                        "source_task_name": task_name,
                        "source_pdf_url": pdf_url,
                        "updated_at": today()
                    }},
                    upsert=True
                )
                print(f"[Context] Contexto XLSX-PDF salvo para {id_usuario_fmt} (Tarefa: {task_name}).")
            except Exception as e:
                print(f"[Context] FALHA ao salvar contexto XLSX-PDF para {id_usuario_fmt}: {e}")
            
            
            if intention == "enigma":
                print(f"[DEBUG-V17] Processando Enigma a partir de texto baixado...")
                
                raw_result = extract_hidden_message(pdf_content_raw)
                
                if not raw_result:
                    result = "Nenhuma mensagem secreta encontrada."
                    final_answer = "Analisei o arquivo, mas não encontrei nenhuma frase secreta."
                else:
                    cleanup_prompt = f"""
                    Minha função Python extraiu a seguinte frase secreta literal: "{raw_result}"
                    
                    Sua tarefa é formatar esta frase para que ela se torne legível em português, seguindo regras MUITO ESTRITAS:
                    
                    1.  **Ordem e Letras:** A ordem das letras e as próprias letras NÃO PODEM MUDAR.
                        (Exemplo: O bloco 'EQ UI PE' deve se tornar 'EQUIPE'. O 'Q' não pode ser inventado. O 'E' inicial não pode virar 'É'.)
                    2.  **Acentuação:** Adicione a acentuação correta onde necessário. (Ex: 'VOCES' -> 'VOCÊS', 'SAO' -> 'SÃO').
                    3.  **Sem Pontuação:** NÃO adicione NENHUMA pontuação (vírgulas, pontos de exclamação, etc.). A saída deve ser "texto liso".
                    4.  **Caixa Alta:** A saída final deve ser TODA EM MAIÚSCULAS.
                    5.  **Palavras Estrangeiras:** PODE HAVER palavras estrangeiras (ex: 'TEAM', 'PROJECT', INNOVATION, etc.), mas mantenha a acentuação e caixa alta conforme as regras acima; se atente ao contexto para descobrir se é uma palavra estrangeira ou brasileira.
                    
                    Frase Bruta: "{raw_result}"
                    Transforme isso em uma frase limpa, em português, toda em maiúsculas, sem pontuação.
                    
                    Exemplo de 'antes' e 'depois':
                    - Antes: "EQ UI PE VO C E S PA SS AR AM"
                    - Depois: "EQUIPE VOCÊS PASSARAM"
                    
                    Responda APENAS com a frase final limpa.
                    """
                    
                    print("[DEBUG-V17-CLEANUP] Chamando IA (Gemini) para limpeza ESTRITA (Lógica Inline)...")
                    
                    cleanup_contents = [Content(role="user", parts=[Part.from_text(cleanup_prompt)])]
                    cleanup_resp = model.generate_content(cleanup_contents, tools=[])
                                    
                    cleaned_result = ""
                    if cleanup_resp.candidates and cleanup_resp.candidates[0].content and cleanup_resp.candidates[0].content.parts:
                        cleaned_result = getattr(cleanup_resp.candidates[0].content.parts[0], "text", "").strip()

                    if not cleaned_result or cleaned_result == raw_result:
                        print("[DEBUG-V17-CLEANUP] IA de limpeza (Inline) falhou. Retornando frase bruta.")
                        result = raw_result
                        final_answer = f"A frase secreta encontrada no PDF da tarefa '{task_name}' é: {raw_result}"
                    else:
                        print(f"[DEBUG-V17-CLEANUP] IA de limpeza (Inline) retornou: {cleaned_result}")
                        result = cleaned_result
                        final_answer = f"A frase secreta encontrada no PDF da tarefa '{task_name}' é: {cleaned_result}"
                tool_steps.append({"call": "solve_pdf_enigma_from_text", "args": {"task": task_name}, "result": result})

            else:
                print(f"[DEBUG-V17] Processando RAG a partir de texto baixado...")
                                
                pdf_content = clean_pdf_text(pdf_content_raw)
                
                rag_prompt = f"""
                Use o CONTEÚDO DO DOCUMENTO abaixo (referente à tarefa '{task_name}') para responder a PERGUNTA DO USUÁRIO.
                ==================== CONTEÚDO DO DOCUMENTO ====================
                {pdf_content[:10000]}
                ===============================================================
                PERGUNTA DO USUÁRIO: {pergunta}
                """                
                data_hoje, (inicio_mes, fim_mes) = iso_date(today()), month_bounds(today())
                system_prompt_filled = SYSTEM_PROMPT.format(
                    nome_usuario=nome_usuario_fmt, 
                    email_usuario=(email_usuario or "email.desconhecido"), 
                    id_usuario=(id_usuario or "id.desconhecido"),
                    data_hoje=data_hoje, inicio_mes=inicio_mes, fim_mes=fim_mes,
                )
                rag_model = init_model(system_prompt_filled)
                rag_resp = rag_model.generate_content([rag_prompt], tools=[])                
                if rag_resp.candidates and rag_resp.candidates[0].content and rag_resp.candidates[0].content.parts:
                    final_answer = getattr(rag_resp.candidates[0].content.parts[0], "text", "Não encontrei a resposta no documento.")
                else:
                    final_answer = "Não encontrei a resposta no documento."
                tool_steps.append({"call": "RAG_on_XLSX_PDF", "args": {"url": pdf_url, "pergunta": pergunta}, "result": final_answer})
        final_answer_fmt = _normalize_answer(final_answer, nome_usuario_fmt)
        response_data = {
            "tipo_resposta": "TEXTO_XLSX",
            "conteudo_texto": final_answer_fmt,
            "dados": tool_steps,
        }
        return JSONResponse(sanitize_doc(response_data))

    except Exception as e:
        raise e
            
@app.post("/pdf/extract-text")
async def pdf_extract_text(file: UploadFile = File(...), _ = Depends(require_api_key)):
    """
    Endpoint utilitário: Envie um PDF e receba o texto completo.
    """
    pdf_bytes = await file.read()
    text = extract_full_pdf_text(pdf_bytes)
    return {"filename": file.filename, "text": text}

@app.post("/pdf/solve-enigma")
async def pdf_solve_enigma(file: UploadFile = File(...), _ = Depends(require_api_key)):
    """
    Endpoint utilitário: Envie um PDF e receba a "frase secreta".
    (NOTA: Este endpoint agora usa a lógica V14, que é confiável)
    """
    pdf_bytes = await file.read()
    text = extract_full_pdf_text(pdf_bytes) 
    message = extract_hidden_message(text)
    return {"filename": file.filename, "message": message or "Nenhuma mensagem encontrada."}

@app.post("/ai/clear-context")
async def clear_chat_context(
    req: ChatRequest,
    _ = Depends(require_api_key)
):
    """
    Limpa o contexto de arquivo (PDF/XLSX) da memória do usuário.
    O Frontend DEVE chamar isso no logout.
    """
    if not req.id_usuario:
        raise HTTPException(status_code=400, detail="id_usuario é obrigatório.")
    
    id_usuario_fmt = req.id_usuario
    nome_usuario_fmt = req.nome_usuario or "você"
    
    try:
        db = mongo()
        result = db[COLL_CONTEXTOS].delete_one({"user_id": id_usuario_fmt})
        
        if result.deleted_count > 0:
            print(f"[Context] Contexto limpo para {id_usuario_fmt}.")
            msg = "Prontinho! Esqueci o arquivo que estávamos analisando. Sobre o que vamos falar agora?"
        else:
            print(f"[Context] Nenhuma contexto para limpar para {id_usuario_fmt}.")
            msg = "Eu não tinha nenhum arquivo na memória, mas estou pronto para o que precisar!"
        
        return JSONResponse({
            "tipo_resposta": "TEXTO",
            "conteudo_texto": _normalize_answer(msg, nome_usuario_fmt),
            "dados": [{"deleted": result.deleted_count}]
        })
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao limpar contexto: {str(e)}")

@app.get("/")
def root():
    return {
        "status": "OK",
        "service": f"{APPLICATION_NAME} (Serviço Unificado de IA e Importação)",
        "model": GEMINI_MODEL_ID,
        "project": PROJECT_ID,
        "location": LOCATION,
        "main_api_target": TASKS_API_BASE,
    }
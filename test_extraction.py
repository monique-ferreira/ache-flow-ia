import os, io, re, time, sys
from typing import List, Optional, Dict, Any, Tuple
from urllib.parse import urlparse, parse_qs, unquote, quote, urljoin

# --- Bibliotecas necessárias ---
# Certifique-se de que estão instaladas no seu ambiente:
# pip install pandas openpyxl requests pdfplumber PyMuPDF
# ---------------------------------
try:
    import requests
    import pandas as pd
    import pdfplumber
    import fitz # PyMuPDF
    from openpyxl import load_workbook
    from openpyxl.utils import range_boundaries
except ImportError:
    print("Erro: Bibliotecas necessárias não encontradas.")
    print("Por favor, instale-as executando:")
    print("pip install pandas openpyxl requests pdfplumber PyMuPDF")
    sys.exit(1)


# =========================
# Constantes (do main.py)
# =========================
TIMEOUT_S = 90
GENERIC_USER_AGENT = "ache-flow-ia/1.0"
PDF_USER_AGENT     = GENERIC_USER_AGENT

# =========================
# Funções de Download (Atualizadas)
# =========================

def fetch_bytes(url: str) -> bytes:
    if not url: raise ValueError("URL ausente")
    u = url.strip()
    if not u.lower().startswith(("http://", "https://")): raise ValueError("URL inválida")
    
    headers = {"User-Agent": GENERIC_USER_AGENT}
    
    is_google_export = "docs.google.com" in u and "export?format=xlsx" in u
    
    if is_google_export:
        with requests.get(u, timeout=TIMEOUT_S, allow_redirects=True, headers=headers, stream=True) as r:
            r.raise_for_status()
            ctype = (r.headers.get("Content-Type") or "").lower()
            if "openxmlformats-officedocument.spreadsheetml.sheet" not in ctype:
                raise ValueError(f"URL do Google não retornou um XLSX. Content-Type: {ctype}")
            content_bytes = io.BytesIO()
            for chunk in r.iter_content(chunk_size=8192):
                content_bytes.write(chunk)
            return content_bytes.getvalue()
    else:
        r = requests.get(u, timeout=TIMEOUT_S, allow_redirects=True, headers=headers)
        r.raise_for_status()
        return r.content

def extract_gsheet_id(url: str) -> Optional[str]:
    if not url: return None
    m = re.search(r"/d/([a-zA-Z0-9_-]{40,})", url)
    if m:
        return m.group(1)
    return None

def _try_download_traced(url: str, timeout: int, user_agent: str):
    headers = {"User-Agent": user_agent, "Accept": "application/pdf, */*"}
    with requests.get(url, timeout=timeout, allow_redirects=True, headers=headers, stream=False) as r:
        status, ctype, final_url, content = r.status_code, (r.headers.get("Content-Type") or "").lower(), r.url, r.content or b""
        return {"status": status, "content_type": ctype, "final_url": final_url, "content": content, "is_pdf_signature": content[:4] == b"%PDF", "size": len(content)}

def normalize_sharepoint_pdf_url(u: str) -> str:
    try:
        pu = urlparse(u)
        if pu.netloc and "sharepoint.com" in pu.netloc and pu.path.endswith("/onedrive.aspx"):
            qs = parse_qs(pu.query or ""); idp = qs.get("id", [None])[0]
            if idp:
                raw_path = unquote(idp);
                if not raw_path.startswith("/"): raw_path = "/" + raw_path
                encoded_path = quote(raw_path, safe="/-_.()~")
                return f"{pu.scheme}://{pu.netloc}{encoded_path}"
    except Exception: pass
    return u

# --- NOVA FUNÇÃO ---
def normalize_gdrive_pdf_url(u: str) -> str:
    """Converte links de visualização do Google Drive em links de download direto."""
    try:
        # Tenta extrair o ID de formatos como /file/d/FILE_ID/view
        m = re.search(r"/file/d/([a-zA-Z0-9_-]{28,})", u)
        if m:
            file_id = m.group(1)
            return f"https://drive.google.com/uc?export=download&id={file_id}"
        
        # Tenta extrair o ID de formatos como /open?id=FILE_ID (links mais antigos)
        m_open = re.search(r"id=([a-zA-Z0-9_-]{28,})", u)
        if m_open and "drive.google.com" in u:
            file_id = m_open.group(1)
            return f"https://drive.google.com/uc?export=download&id={file_id}"
    except Exception:
        pass
    return u
# --- FIM DA NOVA FUNÇÃO ---


# --- FUNÇÃO MODIFICADA ---
def fetch_pdf_bytes(url: str):
    if not url: raise ValueError("URL ausente para PDF")
    u = url.strip()
    if not u.lower().startswith(("http://", "https://")):
        if os.path.exists(url):
            print(f"  [Info] Lendo arquivo PDF local: {url}")
            with open(url, "rb") as f:
                return f.read()
        raise ValueError(f"URL inválida ou arquivo local não encontrado: {url}")
    
    print(f"  [Info] Baixando PDF de: {url[:70]}...")

    # Aplicar normalizadores de URL
    u_norm = u
    if "sharepoint.com" in u:
        u_norm = normalize_sharepoint_pdf_url(u)
    elif "drive.google.com" in u:
        u_norm = normalize_gdrive_pdf_url(u)
    
    def _is_pdf(ctype: str, content: bytes) -> bool: 
        # Links do GDrive podem vir como octet-stream. A assinatura %PDF é a chave.
        return ("pdf" in ctype.lower()) or ("octet-stream" in ctype.lower()) or (content[:4] == b"%PDF")

    variants = [("normalized", u_norm)]
    if u_norm != u: # Adiciona o original como fallback se for diferente
        variants.append(("original", u))

    # Adiciona as variantes do SharePoint SE for um link do sharepoint
    if "sharepoint.com" in u_norm.lower():
        if "download=1" not in u_norm.lower():
            sep = "&" if "?" in u_norm else "?"; variants.append(("download=1", u_norm + f"{sep}download=1"))
        
        pu = urlparse(u_norm); base = f"{pu.scheme}://{pu.netloc}"; src = quote(u_norm, safe=""); variants.append(("download.aspx", f"{base}/_layouts/15/download.aspx?SourceUrl={src}"))
    
    last = None
    for label, cand in variants:
        print(f"    [Debug] Tentando baixar (método: {label}): {cand[:80]}...")
        r = _try_download_traced(cand, TIMEOUT_S, PDF_USER_AGENT); last = r

        # Lida com a página de "aviso de vírus" do Google Drive
        if "drive.google.com" in cand and "text/html" in r["content_type"].lower():
            try:
                print("    [Debug] GDrive retornou HTML. Verificando se é página de confirmação...")
                m_confirm = re.search(r'href="(/uc\?export=download&amp;confirm=[a-zA-Z0-9_&;-]+)"', r["content"].decode('utf-8', errors='ignore'))
                if m_confirm:
                    confirm_url = "https://drive.google.com" + m_confirm.group(1).replace("&amp;", "&")
                    print(f"    [Debug] GDrive precisa de confirmação. Tentando: {confirm_url[:80]}...")
                    # Tenta baixar de novo com o link de confirmação
                    r = _try_download_traced(confirm_url, TIMEOUT_S, PDF_USER_AGENT); last = r
            except Exception as e_parse:
                print(f"    [Debug] Falha ao parsear página de confirmação do GDrive: {e_parse}")

        if r["status"] == 200 and _is_pdf(r["content_type"], r["content"]): 
            print("  [Info] Download concluído.")
            return r["content"]
    
    raise ValueError(f"Não foi possível obter PDF (último status={last['status'] if last else None}, content-type={last['content_type'] if last else None}).")
# --- FIM DA FUNÇÃO MODIFICADA ---


# =========================
# Funções de Extração (do main.py - Sem alteração)
# =========================

def xlsx_bytes_to_dataframe_preserving_hyperlinks(xlsx_bytes: bytes) -> pd.DataFrame:
    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=False); ws = wb.active
    headers: List[str] = [str(cell.value).strip() if cell.value is not None else "" for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=False))]
    hyperlink_map: Dict[str, str] = {}
    for hl in getattr(ws, "_hyperlinks", []) or []:
        try:
            ref = getattr(hl, "ref", None); target = getattr(hl, "target", None) or getattr(hl, "location", None)
            if not ref or not target: continue
            if ":" in ref:
                min_col, min_row, max_col, max_row = range_boundaries(ref)
                for r in range(min_row, max_row + 1):
                    for c in range(min_col, max_col + 1): hyperlink_map[ws.cell(row=r, column=c).coordinate] = target
            else: hyperlink_map[ref] = target
        except Exception: continue
    rows: List[Dict[str, Any]] = []
    col_idx_doc = headers.index("Documento Referência") if "Documento Referência" in headers else -1
    for row in ws.iter_rows(min_row=2, values_only=False):
        if all((c.value is None or str(c.value).strip() == "") for c in row): continue
        record: Dict[str, Any] = {}
        for i, cell in enumerate(row):
            header = headers[i] if i < len(headers) else f"col{i+1}"; val = cell.value
            if i == col_idx_doc:
                url = (getattr(cell.hyperlink, "target", None) if getattr(cell, "hyperlink", None) else None) or hyperlink_map.get(cell.coordinate)
                if not url and isinstance(val, str) and (val.strip().lower().startswith(("http://", "https://", "file://")) or os.path.exists(val.strip())): 
                    url = val.strip()
                record[header] = url or (val if val is not None else "")
            else: record[header] = val if val is not None else ""
        rows.append(record)
    return pd.DataFrame(rows)

def extract_full_pdf_text(pdf_bytes: bytes) -> str:
    text_all = ""
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            text_all = "\n".join((page.extract_text(x_tolerance=3, y_tolerance=3) or "") for page in pdf.pages if page.extract_text())
        if text_all.strip(): 
            return text_all
    except Exception as e:
        print(f"  [Info] PDFPlumber falhou: {e}. Tentando PyMuPDF.")
    
    try:
        text_all_fitz = ""
        with fitz.open(stream=pdf_bytes, filetype="pdf") as doc:
            for page in doc:
                text_all_fitz += page.get_text("text") or ""
        if text_all_fitz.strip():
            return text_all_fitz
    except Exception as e_fitz:
         print(f"  [Erro] PyMuPDF (fitz) também falhou: {e_fitz}")
         return ""
    return ""

def clean_pdf_text(s: str) -> str:
    if not s: return s
    s = re.sub(r"([^\n])\n([^\n])", r"\1 \2", s)
    s = re.sub(r"[ \t\r\f\v]+", " ", s)
    s = re.sub(r"\s+([,;\.\!\?\:\)])", r"\1", s)
    s = re.sub(r"([,;\.\!\?\:])([^\s])", r"\1 \2", s)
    return s.strip()

def _anchor_regex_flex(label: str) -> re.Pattern:
    m = re.search(r"(?i)texto\.?(\d+)", label or "");
    if not m: return re.compile(r"(?!)")
    num = m.group(1); return re.compile(rf"(?i)\bTexto\.?{re.escape(num)}\.?\b[:\-]?\s*")

def extract_after_anchor_from_pdf(text_all: str, anchor_label: str, max_chars: int = 4000) -> str:
    if not text_all.strip(): return ""
    rx = _anchor_regex_flex(anchor_label); m = rx.search(text_all)
    if not m: 
        print(f"  [Aviso] Âncora '{anchor_label}' não encontrada no PDF.")
        return ""
    start = m.end()
    next_m = re.search(r"(?i)\bTexto\.?\d+\.?\b", text_all[start:])
    end = start + next_m.start() if next_m else len(text_all)
    return text_all[start:end].strip()[:max_chars].strip()

def resolve_descricao_pdf(row) -> str:
    como, docrf = str(row.get("Como Fazer") or "").strip(), str(row.get("Documento Referência") or "").strip()
    
    if not como or not docrf or not re.search(r"(?i)\b((?:Doc\.?\s*)?(Texto\.?\d+))\b\.?", como): 
        return como
    
    try: 
        pdf_bytes = fetch_pdf_bytes(docrf)
        full_pdf_text = extract_full_pdf_text(pdf_bytes)
        if not full_pdf_text.strip():
            print(f"  [Aviso] PDF {docrf} não retornou texto. Usando '{como}' como fallback.")
            return como
            
    except Exception as e:
        print(f"  [Erro] Falha ao baixar/processar PDF {docrf}: {e}. Usando '{como}' como fallback.")
        return como

    def _repl(m: re.Match) -> str:
        full_token, anchor = m.group(1), m.group(2)
        extracted = clean_pdf_text(extract_after_anchor_from_pdf(full_pdf_text, anchor))
        return extracted if extracted else full_token

    return re.sub(r"(?i)\b((?:Doc\.?\s*)?(Texto\.?\d+))\b\.?", _repl, como)

# =========================
# Lógica Principal do Teste
# =========================
def run_test():
    # --- Configure o LINK da sua planilha aqui ---
    xlsx_url_to_test = r"https://docs.google.com/spreadsheets/d/1Pv_jFkiYVOm9QEDfwZgaacLXPD1MToCvspjczhKZoOI/edit?usp=sharing"
    # ----------------------------------------------------

    if xlsx_url_to_test == "COLE_A_URL_DA_PLANILHA_AQUI":
        print("ERRO: Por favor, edite o script 'test_extraction.py'")
        print("      e cole a URL da sua planilha na variável 'xlsx_url_to_test'.")
        return

    print(f"Iniciando teste de extração para URL: {xlsx_url_to_test}\n")
    
    try:
        if "docs.google.com/spreadsheets" in xlsx_url_to_test:
            gsheet_id = extract_gsheet_id(xlsx_url_to_test)
            if not gsheet_id: raise ValueError("URL do Google Sheets inválida ou ID não extraído.")
            download_url = f"https://docs.google.com/spreadsheets/d/{gsheet_id}/export?format=xlsx"
            print(f"  [Info] URL de Google Sheet detectada. Baixando de: {download_url}")
        else:
            download_url = xlsx_url_to_test
            print(f"  [Info] Baixando XLSX de: {download_url[:70]}...")

        file_bytes = fetch_bytes(download_url)
        print(f"  [Info] XLSX baixado com sucesso ({len(file_bytes)} bytes).")

    except Exception as e:
        print(f"ERRO ao baixar ou processar a URL da planilha: {e}")
        return

    df = xlsx_bytes_to_dataframe_preserving_hyperlinks(file_bytes)
    
    if "Como Fazer" not in df.columns or "Documento Referência" not in df.columns:
        print(f"ERRO: Colunas 'Como Fazer' ou 'Documento Referência' não encontradas no XLSX.")
        print(f"Colunas encontradas: {list(df.columns)}")
        return

    print("="*80)
    
    for index, row in df.iterrows():
        nome_tarefa = str(row.get("Nome") or f"Linha {index + 2}")
        como_fazer = str(row.get("Como Fazer") or "").strip()
        doc_ref_url = str(row.get("Documento Referência") or "").strip()
        
        print(f"\n[TAREFA: {nome_tarefa}]")
        print(f"  Como Fazer (Original): {como_fazer}")
        print(f"  Documento Ref (URL): {doc_ref_url}")
        
        descricao_final = resolve_descricao_pdf(row)
        
        print(f"  Descrição (Final): {descricao_final}")
        print("-"*80)

    print("\nTeste concluído.")


if __name__ == "__main__":
    run_test()